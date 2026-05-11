"""
Staggered DiD (Callaway & Sant'Anna 2021)
Treatment: first year NEW DC capacity (added since 2016) exceeds threshold.
Control: never-treated states.

Thresholds: 500MW, 1GW
Y variables: employment (raw 5182, cleaned direct, cleaned × multiplier)
             + construction (NAICS 23, 2362, 5415)

Output: 06_staggered_did_analysis.xlsx
  Sheet 1: Overall ATT
  Sheet 2: Event Study (dynamic ATTs by relative time)
  Sheet 3: Treatment Cohorts
  Sheet 4: Event Study Plot Data (wide format for charting)
"""

import pandas as pd
import numpy as np
import warnings
import io
import sys
from pathlib import Path
from collections import Counter
from scipy.stats import norm
from csdid import att_gt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

_DIR = Path(__file__).resolve().parent / "employment"


class SuppressPrint:
    def __enter__(self):
        self._orig = sys.stdout
        sys.stdout = io.StringIO()
        return self

    def __exit__(self, *args):
        sys.stdout = self._orig


# ============================================================
# 1. Load data
# ============================================================
qwi = pd.read_csv(_DIR / 'qwi_all_naics_annual.csv', dtype={'naics': str})
dc = pd.read_csv(_DIR / 'dc_facilities_by_state_year.csv')
mult = pd.read_csv(_DIR / 'pwc_multipliers.csv')

STATE_ABBR_MAP = {
    'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA',
    'Colorado':'CO','Connecticut':'CT','Delaware':'DE','District of Columbia':'DC',
    'Florida':'FL','Georgia':'GA','Hawaii':'HI','Idaho':'ID','Illinois':'IL',
    'Indiana':'IN','Iowa':'IA','Kansas':'KS','Kentucky':'KY','Louisiana':'LA',
    'Maine':'ME','Maryland':'MD','Massachusetts':'MA','Michigan':'MI','Minnesota':'MN',
    'Mississippi':'MS','Missouri':'MO','Montana':'MT','Nebraska':'NE','Nevada':'NV',
    'New Hampshire':'NH','New Jersey':'NJ','New Mexico':'NM','New York':'NY',
    'North Carolina':'NC','North Dakota':'ND','Ohio':'OH','Oklahoma':'OK','Oregon':'OR',
    'Pennsylvania':'PA','Rhode Island':'RI','South Carolina':'SC','South Dakota':'SD',
    'Tennessee':'TN','Texas':'TX','Utah':'UT','Vermont':'VT','Virginia':'VA',
    'Washington':'WA','West Virginia':'WV','Wisconsin':'WI','Wyoming':'WY'
}
ABBR_TO_NAME = {v: k for k, v in STATE_ABBR_MAP.items()}
YEARS = list(range(2016, 2025))

# ============================================================
# 2. Spillover decomposition
# ============================================================
NAT_MULT = mult['emp_total_no_spillover_avg'].sum() / mult['emp_direct_avg'].sum()
mult['spillover_dc'] = mult['emp_spillover_avg'] / NAT_MULT
mult['local_share'] = mult['emp_direct_avg'] / (mult['emp_direct_avg'] + mult['spillover_dc'])

local_share_dict = dict(zip(mult['state'], mult['local_share']))
local_mult_dict = dict(zip(mult['state'], mult['emp_multiplier_local']))

# ============================================================
# 3. Build panels
# ============================================================
dc_states = dc[dc['state_abbr'] != 'US'].copy()
dc_panel = []
for _, row in dc_states.iterrows():
    for yr in YEARS:
        dc_panel.append({
            'state_abbr': row['state_abbr'], 'year': yr,
            'dc_capacity_GW': row[f'MW_{yr}'] / 1000
        })
dc_panel = pd.DataFrame(dc_panel)
state_ids = {s: i + 1 for i, s in enumerate(sorted(dc_panel['state_abbr'].unique()))}

qwi['state_abbr'] = qwi['state_name'].map(STATE_ABBR_MAP)

# Employment panel
q5182 = qwi[qwi['naics'] == '5182'][['state_abbr', 'year', 'Emp_annual_avg']].rename(
    columns={'Emp_annual_avg': 'emp_5182'})
emp_panel = q5182.merge(dc_panel[['state_abbr', 'year']], on=['state_abbr', 'year'], how='inner')
emp_panel['state_id'] = emp_panel['state_abbr'].map(state_ids)
emp_panel['state_name'] = emp_panel['state_abbr'].map(ABBR_TO_NAME)
emp_panel['local_share'] = emp_panel['state_name'].map(local_share_dict)
emp_panel['local_multiplier'] = emp_panel['state_name'].map(local_mult_dict)
emp_panel['emp_cleaned_direct'] = emp_panel['emp_5182'] * emp_panel['local_share']
emp_panel['emp_cleaned_multiplied'] = emp_panel['emp_cleaned_direct'] * emp_panel['local_multiplier']

# Construction panels
const_panels = {}
for naics in ['23', '2362', '5415']:
    q = qwi[qwi['naics'] == naics][['state_abbr', 'year', 'Emp_annual_avg']].rename(
        columns={'Emp_annual_avg': f'emp_{naics}'})
    cp = q.merge(dc_panel[['state_abbr', 'year']], on=['state_abbr', 'year'], how='inner')
    cp['state_id'] = cp['state_abbr'].map(state_ids)
    const_panels[naics] = cp


# ============================================================
# 4. Treatment assignment
# ============================================================
def get_treatment_years(threshold_gw):
    treat = {}
    for abbr in dc_panel['state_abbr'].unique():
        st = dc_panel[dc_panel['state_abbr'] == abbr].sort_values('year')
        cap_2016 = st[st['year'] == 2016]['dc_capacity_GW'].values[0]
        crossed = st[(st['dc_capacity_GW'] - cap_2016 >= threshold_gw) & (st['year'] > 2016)]
        treat[abbr] = crossed['year'].iloc[0] if len(crossed) > 0 else 0
    return treat


# ============================================================
# 5. Run all specs
# ============================================================
ALL_SPECS = [
    ('Employment', 'emp_5182', 'Raw NAICS 5182', emp_panel),
    ('Employment', 'emp_cleaned_direct', 'Cleaned Direct', emp_panel),
    ('Employment', 'emp_cleaned_multiplied', 'Cleaned × Multiplier', emp_panel),
    ('Construction', 'emp_23', 'NAICS 23 (Construction)', const_panels['23']),
    ('Construction', 'emp_2362', 'NAICS 2362 (Nonresidential)', const_panels['2362']),
    ('Construction', 'emp_5415', 'NAICS 5415 (Computer Systems)', const_panels['5415']),
]

all_simple = []
all_dynamic = []
all_cohorts = []

for threshold_gw, threshold_label in [(0.5, '500MW'), (1.0, '1GW')]:
    treat_years = get_treatment_years(threshold_gw)
    treated_states = {k: v for k, v in treat_years.items() if v > 0}
    control_states = {k: v for k, v in treat_years.items() if v == 0}
    n_t, n_c = len(treated_states), len(control_states)

    cohorts = Counter(treated_states.values())
    for yr in sorted(cohorts):
        sts = sorted([k for k, v in treated_states.items() if v == yr])
        all_cohorts.append({
            'threshold': threshold_label, 'cohort_year': yr,
            'n_states': len(sts), 'states': ', '.join(sts)
        })

    for category, y_col, y_label, base_panel in ALL_SPECS:
        df = base_panel[['state_id', 'state_abbr', 'year', y_col]].copy()
        df['treat_year'] = df['state_abbr'].map(treat_years)
        df = df.dropna(subset=[y_col])

        try:
            with SuppressPrint():
                model = att_gt.ATTgt(
                    yname=y_col, tname='year', idname='state_id',
                    gname='treat_year', data=df, control_group='nevertreated'
                )
                model.fit()
                agg_s = model.aggte(typec='simple')
                agg_d = model.aggte(typec='dynamic')

            att = float(agg_s.atte['overall_att'])
            se = float(agg_s.atte['overall_se'].flat[0])
            p = 2 * (1 - norm.cdf(abs(att / se)))

            all_simple.append({
                'threshold': threshold_label, 'category': category, 'y_var': y_label,
                'ATT': att, 'se': se, 'ci_low': att - 1.96 * se, 'ci_high': att + 1.96 * se,
                'p_approx': p, 'n_treated': n_t, 'n_control': n_c,
            })

            egt = agg_d.atte['egt']
            att_egt = agg_d.atte['att_egt']
            se_egt = agg_d.atte['se_egt']
            if se_egt.ndim == 2:
                se_egt = se_egt[0]

            for e, a, s in zip(egt, att_egt, se_egt):
                all_dynamic.append({
                    'threshold': threshold_label, 'category': category, 'y_var': y_label,
                    'event_time': int(e), 'ATT': float(a), 'se': float(s),
                    'ci_low': float(a - 1.96 * s), 'ci_high': float(a + 1.96 * s),
                })

            sig = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else ''
            print(f"OK  {threshold_label:5s}  {y_label:30s}  ATT={att:>10,.1f}  SE={se:>8,.1f}  p={p:.4f}{sig}")
        except Exception as ex:
            print(f"ERR {threshold_label:5s}  {y_label:30s}  {ex}")

print(f"\nTotal: {len(all_simple)} overall, {len(all_dynamic)} event-study, {len(all_cohorts)} cohorts")

# ============================================================
# 6. Write Excel
# ============================================================
thin = Side(style='thin', color='000000')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
POS_FILL = PatternFill("solid", fgColor="C6EFCE")
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")

wb = Workbook()
wb.remove(wb.active)

# --- Sheet 1: Overall ATT ---
ws = wb.create_sheet("Overall ATT")
ws.sheet_properties.tabColor = "4472C4"
ws.freeze_panes = "A3"
hf = Font(bold=True, color="FFFFFF", size=11)

ws.merge_cells('A1:J1')
c = ws['A1']
c.value = "Staggered DiD (Callaway & Sant'Anna 2021) — Overall ATT"
c.font = hf
c.fill = PatternFill("solid", fgColor="4472C4")
c.alignment = Alignment(horizontal='center')

headers = ['Threshold', 'Category', 'Y Variable', 'ATT', 'SE', 'CI95 Lower',
           'CI95 Upper', 'p-value', 'Treated', 'Control']
hdr_fill = PatternFill("solid", fgColor="D6E4F0")
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=ci, value=h)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
    cell.fill = hdr_fill

fmts = [None, None, None, '#,##0.0', '#,##0.0', '#,##0.0', '#,##0.0', '0.0000', None, None]
for idx, r in enumerate(all_simple):
    rn = idx + 3
    vals = [r['threshold'], r['category'], r['y_var'], r['ATT'], r['se'],
            r['ci_low'], r['ci_high'], r['p_approx'], r['n_treated'], r['n_control']]
    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws.cell(row=rn, column=ci, value=v)
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
    if r['p_approx'] < 0.05:
        fill = POS_FILL if r['ATT'] > 0 else NEG_FILL
        for ci in [4, 5, 6, 7, 8]:
            ws.cell(row=rn, column=ci).fill = fill

sr = len(all_simple) + 4
ws.cell(row=sr, column=1, value="Notes").font = Font(bold=True, size=11)
for i, (l, v) in enumerate([
    ("Model", "Callaway & Sant'Anna (2021) Doubly Robust DiD"),
    ("Treatment", "First year NEW DC capacity (since 2016) > threshold"),
    ("Control", "Never-treated states (never exceeded threshold by 2024)"),
    ("SE", "Bootstrapped (1000 iterations)"),
    ("Period", "2016–2024"),
    ("ATT", "Average treatment effect on the treated"),
]):
    ws.cell(row=sr + 1 + i, column=1, value=l).font = Font(bold=True if l else False, size=10)
    ws.cell(row=sr + 1 + i, column=3, value=v)

for ci, w in enumerate([10, 14, 28, 12, 10, 12, 12, 10, 8, 8], 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# --- Sheet 2: Event Study ---
ws2 = wb.create_sheet("Event Study")
ws2.sheet_properties.tabColor = "7030A0"
ws2.freeze_panes = "A2"
headers2 = ['Threshold', 'Category', 'Y Variable', 'Event Time', 'ATT', 'SE',
            'CI95 Lower', 'CI95 Upper', 'Sig']
for ci, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
    cell.fill = PatternFill("solid", fgColor="E8D5F5")

fmts2 = [None, None, None, None, '#,##0.0', '#,##0.0', '#,##0.0', '#,##0.0', None]
for idx, r in enumerate(all_dynamic):
    rn = idx + 2
    se_v = r['se']
    att_v = r['ATT']
    sig = (abs(att_v / se_v) > 1.96) if se_v > 0 else False
    vals = [r['threshold'], r['category'], r['y_var'], r['event_time'],
            att_v, se_v, r['ci_low'], r['ci_high'], 'Yes' if sig else '']
    for ci, (v, fmt) in enumerate(zip(vals, fmts2), 1):
        cell = ws2.cell(row=rn, column=ci, value=v)
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
    if r['event_time'] < 0 and sig:
        for ci in [5, 6, 9]:
            ws2.cell(row=rn, column=ci).fill = NEG_FILL
    elif r['event_time'] >= 0 and sig:
        fill = POS_FILL if att_v > 0 else NEG_FILL
        for ci in [5, 6, 9]:
            ws2.cell(row=rn, column=ci).fill = fill

for ci, w in enumerate([10, 14, 28, 10, 12, 10, 12, 12, 6], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

# --- Sheet 3: Treatment Cohorts ---
ws3 = wb.create_sheet("Treatment Cohorts")
ws3.sheet_properties.tabColor = "548235"
ws3.freeze_panes = "A2"
for ci, h in enumerate(['Threshold', 'Cohort Year', '# States', 'States'], 1):
    cell = ws3.cell(row=1, column=ci, value=h)
    cell.font = Font(bold=True, size=10)
    cell.border = BORDER
    cell.fill = PatternFill("solid", fgColor="E2EFDA")
for idx, r in enumerate(all_cohorts):
    rn = idx + 2
    for ci, v in enumerate([r['threshold'], r['cohort_year'], r['n_states'], r['states']], 1):
        ws3.cell(row=rn, column=ci, value=v).border = BORDER
for ci, w in enumerate([10, 12, 10, 60], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# --- Sheet 4: Event Study Plot Data ---
ws4 = wb.create_sheet("Event Study Plot Data")
ws4.sheet_properties.tabColor = "BF8F00"
key_specs = [
    ('500MW', 'Cleaned Direct'),
    ('500MW', 'NAICS 2362 (Nonresidential)'),
    ('1GW', 'Cleaned Direct'),
    ('1GW', 'NAICS 2362 (Nonresidential)'),
]
col_idx = 1
for threshold, y_var in key_specs:
    sub = sorted(
        [r for r in all_dynamic if r['threshold'] == threshold and r['y_var'] == y_var],
        key=lambda x: x['event_time']
    )
    if not sub:
        continue
    ws4.cell(row=1, column=col_idx, value=f"{y_var} ({threshold})").font = Font(bold=True, size=10)
    for ci, h in enumerate(['e', 'ATT', 'CI_lo', 'CI_hi']):
        c = ws4.cell(row=2, column=col_idx + ci, value=h)
        c.font = Font(bold=True, size=9)
        c.border = BORDER
    for i, r in enumerate(sub):
        rn = i + 3
        for ci, v in enumerate([r['event_time'], r['ATT'], r['ci_low'], r['ci_high']]):
            c = ws4.cell(row=rn, column=col_idx + ci, value=v)
            c.border = BORDER
            if ci > 0:
                c.number_format = '#,##0.0'
    col_idx += 5

wb.save(_DIR / '06_staggered_did_analysis.xlsx')
print(f"\nSaved 06_staggered_did_analysis.xlsx")
print(f"Sheets: {wb.sheetnames}")
