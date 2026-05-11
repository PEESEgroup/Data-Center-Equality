"""
OLS: Employment = α + β × DC_capacity(GW) + ε
State-level, 2016-2024.

Four employment specifications:
  1. Raw NAICS 5182 (observed direct)
  2. Raw NAICS 5182 × local_multiplier + spillover (double-counting issue)
  3. Cleaned direct = NAICS 5182 × local_share (removing spillover-driven 5182 jobs)
  4. Cleaned direct × local_multiplier (total local impact, no double-counting)

Spillover decomposition method:
  - National multiplier = Σ(total_no_spillover) / Σ(direct) ≈ 5.43
  - DC industry share of total = 1/national_multiplier ≈ 18%
  - spillover_dc_i = PwC_spillover_i / national_multiplier
  - local_share_i = PwC_direct_i / (PwC_direct_i + spillover_dc_i)

Outputs: 02_ols_employment_dc_capacity.xlsx
  Sheet 1: OLS by Capacity (raw)
  Sheet 2: Panel Data
  Sheet 3: Cleaned OLS by Capacity
  Sheet 4: Panel FE (State + Year fixed effects, clustered SE)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from scipy import stats
from linearmodels.panel import PanelOLS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 1. Load data
# ============================================================
_DIR = Path(__file__).resolve().parent / "employment"

qwi = pd.read_csv(_DIR / 'qwi_all_naics_annual.csv', dtype={'naics': str})
dc = pd.read_csv(_DIR / 'dc_facilities_by_state_year.csv')
mult = pd.read_csv(_DIR / 'pwc_multipliers.csv')


STATE_ABBR_MAP = {
    'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA',
    'Colorado':'CO','Connecticut':'CT','Delaware':'DE','District of Columbia':'DC',
    'Florida':'FL','Georgia':'GA','Hawaii':'HI','Idaho':'ID','Illinois':'IL',
    'Indiana':'IN','Iowa':'IA','Kansas':'KS','Kentucky':'KY','Louisiana':'LA',
    'Maine':'ME','Maryland':'MD','Massachusetts':'MA','Michigan':'MI','Minnesota':'MN',
    'Mississippi':'MS','Missouri':'MO','Montana':'MT','Nebraska':'NE','Nevada':'NV',
    'New Hampshire':'NH','New Jersey':'NJ','New Mexico':'NM','New York':'NY',
    'North Carolina':'NC','North Dakota':'ND','Ohio':'OH','Oklahoma':'OK','Oregon':'OR',
    'Pennsylvania':'PA','Rhode Island':'RI','South Carolina':'SC','South Dakota':'SD',
    'Tennessee':'TN','Texas':'TX','Utah':'UT','Vermont':'VT','Virginia':'VA',
    'Washington':'WA','West Virginia':'WV','Wisconsin':'WI','Wyoming':'WY'
}

YEARS = list(range(2016, 2025))

# ============================================================
# 2. National multiplier and spillover decomposition
# ============================================================
NAT_MULT = mult['emp_total_no_spillover_avg'].sum() / mult['emp_direct_avg'].sum()  # ≈ 5.43

mult['spillover_dc'] = mult['emp_spillover_avg'] / NAT_MULT
mult['local_share'] = mult['emp_direct_avg'] / (mult['emp_direct_avg'] + mult['spillover_dc'])

local_share_dict = dict(zip(mult['state'], mult['local_share']))
local_mult_dict = dict(zip(mult['state'], mult['emp_multiplier_local']))
spillover_rate_dict = dict(zip(mult['state'], mult['emp_spillover_rate']))

# ============================================================
# 3. Build panel: state × year
# ============================================================
qwi_5182 = qwi[qwi['naics'] == '5182'].copy()
qwi_5182['state_abbr'] = qwi_5182['state_name'].map(STATE_ABBR_MAP)

dc_states = dc[dc['state_abbr'] != 'US'].copy()
dc_panel = []
for _, row in dc_states.iterrows():
    for yr in YEARS:
        dc_panel.append({
            'state_abbr': row['state_abbr'],
            'year': yr,
            'dc_count': row[f'count_{yr}'],
            'dc_capacity_GW': row[f'MW_{yr}'] / 1000
        })
dc_panel = pd.DataFrame(dc_panel)

direct_by_year = qwi_5182.groupby('year')['Emp_annual_avg'].sum().to_dict()

panel = qwi_5182[['state_abbr', 'state_name', 'year', 'Emp_annual_avg']].merge(
    dc_panel, on=['state_abbr', 'year'], how='inner'
)
panel = panel.rename(columns={'Emp_annual_avg': 'emp_5182'})

# Raw multiplied (spec 2, has double-counting)
def calc_raw_multiplied(row):
    st = row['state_name']
    if st not in local_mult_dict:
        return np.nan
    direct_i = row['emp_5182']
    other_direct = direct_by_year.get(row['year'], 0) - direct_i
    return direct_i * local_mult_dict[st] + other_direct * spillover_rate_dict[st]

panel['emp_raw_multiplied'] = panel.apply(calc_raw_multiplied, axis=1)

# Cleaned direct (spec 3) and cleaned multiplied (spec 4)
panel['local_share'] = panel['state_name'].map(local_share_dict)
panel['local_multiplier'] = panel['state_name'].map(local_mult_dict)
panel['emp_cleaned_direct'] = panel['emp_5182'] * panel['local_share']
panel['emp_cleaned_multiplied'] = panel['emp_cleaned_direct'] * panel['local_multiplier']

# ============================================================
# 4. OLS regression
# ============================================================
def run_ols(panel_df, x_col, y_direct_col, y_mult_col, extra_cols=None):
    results = []
    for st_abbr in sorted(panel_df['state_abbr'].unique()):
        df_st = panel_df[panel_df['state_abbr'] == st_abbr].dropna(subset=[y_direct_col, x_col])
        st_name = df_st['state_name'].iloc[0]
        n = len(df_st)
        x = df_st[x_col].values.astype(float)

        base = {'state_abbr': st_abbr, 'state_name': st_name, 'n_obs': n,
                'x_min': x.min() if len(x) > 0 else np.nan,
                'x_max': x.max() if len(x) > 0 else np.nan}

        if extra_cols:
            for ec in extra_cols:
                base[ec] = df_st[ec].iloc[0] if ec in df_st.columns else np.nan

        if n < 3 or np.std(x) < 1e-10:
            for pf in ['direct_', 'mult_']:
                for f in ['slope','se','ci95_lower','ci95_upper','intercept','r_squared','p_value']:
                    base[pf+f] = np.nan
            results.append(base)
            continue

        for pf, yc in [('direct_', y_direct_col), ('mult_', y_mult_col)]:
            y = df_st[yc].values
            sl, ic, r, p, se = stats.linregress(x, y)
            ci = 1.96 * se
            base.update({pf+'slope': sl, pf+'se': se, pf+'ci95_lower': sl-ci,
                         pf+'ci95_upper': sl+ci, pf+'intercept': ic,
                         pf+'r_squared': r**2, pf+'p_value': p})
        results.append(base)
    return pd.DataFrame(results)

# Raw regressions (specs 1 & 2)
res_raw_cap = run_ols(panel, 'dc_capacity_GW', 'emp_5182', 'emp_raw_multiplied')

# Cleaned regressions (specs 3 & 4)
res_cln_cap = run_ols(panel, 'dc_capacity_GW', 'emp_cleaned_direct', 'emp_cleaned_multiplied',
                       extra_cols=['local_share', 'local_multiplier'])

# ============================================================
# 5. Write Excel
# ============================================================
thin = Side(style='thin', color='000000')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
SIG_FILL = PatternFill("solid", fgColor="C6EFCE")
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")


def write_raw_sheet(wb, name, tab_color, res_df, x_label, x_unit, x_fmt,
                    hdr1, hdr2, sub1, sub2, slope_fmt):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A3"
    hf = Font(bold=True, color="FFFFFF", size=11)

    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = "State Info"; c.font = hf
    c.fill = PatternFill("solid", fgColor=hdr1); c.alignment = Alignment(horizontal='center')
    ws.merge_cells('F1:L1')
    c = ws['F1']; c.value = f"NAICS 5182 OLS (by {x_label})"; c.font = hf
    c.fill = PatternFill("solid", fgColor=hdr2); c.alignment = Alignment(horizontal='center')
    ws.merge_cells('M1:S1')
    c = ws['M1']; c.value = f"Raw Multiplier+Spillover OLS (by {x_label})"; c.font = hf
    c.fill = PatternFill("solid", fgColor="548235"); c.alignment = Alignment(horizontal='center')

    cols = [('A','State'),('B','State Name'),('C','N'),
            ('D',f'{x_label}\nMin'),('E',f'{x_label}\nMax'),
            ('F',f'β\n(jobs/{x_unit})'),('G','SE'),('H','CI95\nLower'),('I','CI95\nUpper'),
            ('J','Intercept'),('K','R²'),('L','p-value'),
            ('M',f'β\n(jobs/{x_unit})'),('N','SE'),('O','CI95\nLower'),('P','CI95\nUpper'),
            ('Q','Intercept'),('R','R²'),('S','p-value')]

    fi = PatternFill("solid", fgColor=sub1)
    fd = PatternFill("solid", fgColor=sub2)
    fm = PatternFill("solid", fgColor="E2EFDA")
    for cl, t in cols:
        cell = ws[f'{cl}2']; cell.value = t; cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        cell.fill = fi if cl in 'ABCDE' else (fd if cl in 'FGHIJKL' else fm)
    ws.row_dimensions[2].height = 35

    for idx, (_, r) in enumerate(res_df.sort_values('direct_slope', ascending=False, na_position='last').iterrows()):
        rn = idx + 3
        ws.cell(row=rn,column=1,value=r['state_abbr']).border = BORDER
        ws.cell(row=rn,column=2,value=r['state_name']).border = BORDER
        c = ws.cell(row=rn,column=3,value=int(r['n_obs'])); c.border = BORDER; c.alignment = Alignment(horizontal='center')
        for ci, v in [(4,r['x_min']),(5,r['x_max'])]:
            c = ws.cell(row=rn,column=ci,value=v if not pd.isna(v) else None); c.number_format = x_fmt; c.border = BORDER
        for pf, cs in [('direct_',6),('mult_',13)]:
            for j, (f, fmt) in enumerate(zip(['slope','se','ci95_lower','ci95_upper','intercept','r_squared','p_value'],
                                              [slope_fmt,slope_fmt,slope_fmt,slope_fmt,'#,##0','0.000','0.0000'])):
                v = r[pf+f]
                c = ws.cell(row=rn,column=cs+j,value=v if not pd.isna(v) else None); c.number_format = fmt; c.border = BORDER
        if not pd.isna(r['direct_p_value']) and r['direct_p_value'] < 0.05:
            fl = SIG_FILL if r['direct_slope'] > 0 else NEG_FILL
            for ci in [6,7,8,9,12]: ws.cell(row=rn,column=ci).fill = fl
        if not pd.isna(r['mult_p_value']) and r['mult_p_value'] < 0.05:
            fl = SIG_FILL if r['mult_slope'] > 0 else NEG_FILL
            for ci in [13,14,15,16,19]: ws.cell(row=rn,column=ci).fill = fl

    sr = rn + 2
    ws.cell(row=sr,column=1,value="Summary").font = Font(bold=True, size=11)
    valid = res_df.dropna(subset=['direct_slope'])
    for i, (l, v) in enumerate([
        ("Valid regressions", f"{len(valid)} / {len(res_df)}"),
        (f"Median β NAICS 5182 (jobs/{x_unit})", f"{valid['direct_slope'].median():,.1f}"),
        (f"Median β raw multiplied (jobs/{x_unit})", f"{valid['mult_slope'].median():,.1f}"),
        ("p<0.05 (NAICS 5182)", f"{(valid['direct_p_value']<0.05).sum()} / {len(valid)}"),
        ("p<0.05 (raw multiplied)", f"{(valid['mult_p_value']<0.05).sum()} / {len(valid)}"),
        ("",""),("Model", f"Employment = α + β × {x_label} + ε"),
        ("Note", "Raw multiplied has double-counting (see Cleaned sheets)"),
    ]):
        ws.cell(row=sr+1+i,column=1,value=l).font = Font(bold=True if l else False, size=10)
        ws.cell(row=sr+1+i,column=3,value=v)

    for col, w in {'A':7,'B':22,'C':5,'D':9,'E':9,'F':13,'G':10,'H':13,'I':13,'J':12,'K':8,'L':9,
                   'M':13,'N':10,'O':13,'P':13,'Q':12,'R':8,'S':9}.items():
        ws.column_dimensions[col].width = w


def write_cleaned_sheet(wb, name, tab_color, res_df, x_label, x_unit, x_fmt,
                        hdr1, hdr2, slope_fmt):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A3"
    hf = Font(bold=True, color="FFFFFF", size=11)

    ws.merge_cells('A1:G1')
    c = ws['A1']; c.value = "State Info"; c.font = hf
    c.fill = PatternFill("solid", fgColor=hdr1); c.alignment = Alignment(horizontal='center')
    ws.merge_cells('H1:N1')
    c = ws['H1']; c.value = f"Cleaned Direct OLS (by {x_label})"; c.font = hf
    c.fill = PatternFill("solid", fgColor=hdr2); c.alignment = Alignment(horizontal='center')
    ws.merge_cells('O1:U1')
    c = ws['O1']; c.value = f"Cleaned × Local Multiplier (by {x_label})"; c.font = hf
    c.fill = PatternFill("solid", fgColor="548235"); c.alignment = Alignment(horizontal='center')

    cols = [('A','State'),('B','State Name'),('C','N'),
            ('D','Local\nShare'),('E','Local\nMult'),
            ('F',f'{x_label}\nMin'),('G',f'{x_label}\nMax'),
            ('H',f'β\n(jobs/{x_unit})'),('I','SE'),('J','CI95\nLower'),('K','CI95\nUpper'),
            ('L','Intercept'),('M','R²'),('N','p-value'),
            ('O',f'β\n(jobs/{x_unit})'),('P','SE'),('Q','CI95\nLower'),('R','CI95\nUpper'),
            ('S','Intercept'),('T','R²'),('U','p-value')]

    fi = PatternFill("solid", fgColor="FFF2CC")
    fd = PatternFill("solid", fgColor="D6E4F0")
    fm = PatternFill("solid", fgColor="E2EFDA")
    for cl, t in cols:
        cell = ws[f'{cl}2']; cell.value = t; cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        cell.fill = fi if cl in 'ABCDEFG' else (fd if cl in 'HIJKLMN' else fm)
    ws.row_dimensions[2].height = 35

    for idx, (_, r) in enumerate(res_df.sort_values('direct_slope', ascending=False, na_position='last').iterrows()):
        rn = idx + 3
        ws.cell(row=rn,column=1,value=r['state_abbr']).border = BORDER
        ws.cell(row=rn,column=2,value=r['state_name']).border = BORDER
        c = ws.cell(row=rn,column=3,value=int(r['n_obs'])); c.border = BORDER; c.alignment = Alignment(horizontal='center')
        c = ws.cell(row=rn,column=4,value=r.get('local_share')); c.number_format = '0.0%'; c.border = BORDER
        c = ws.cell(row=rn,column=5,value=r.get('local_multiplier')); c.number_format = '0.00'; c.border = BORDER
        for ci, v in [(6,r['x_min']),(7,r['x_max'])]:
            c = ws.cell(row=rn,column=ci,value=v if not pd.isna(v) else None); c.number_format = x_fmt; c.border = BORDER
        for pf, cs in [('direct_',8),('mult_',15)]:
            for j, (f, fmt) in enumerate(zip(['slope','se','ci95_lower','ci95_upper','intercept','r_squared','p_value'],
                                              [slope_fmt,slope_fmt,slope_fmt,slope_fmt,'#,##0','0.000','0.0000'])):
                v = r[pf+f]
                c = ws.cell(row=rn,column=cs+j,value=v if not pd.isna(v) else None); c.number_format = fmt; c.border = BORDER
        if not pd.isna(r['direct_p_value']) and r['direct_p_value'] < 0.05:
            fl = SIG_FILL if r['direct_slope'] > 0 else NEG_FILL
            for ci in [8,9,10,11,14]: ws.cell(row=rn,column=ci).fill = fl
        if not pd.isna(r['mult_p_value']) and r['mult_p_value'] < 0.05:
            fl = SIG_FILL if r['mult_slope'] > 0 else NEG_FILL
            for ci in [15,16,17,18,21]: ws.cell(row=rn,column=ci).fill = fl

    sr = rn + 2
    ws.cell(row=sr,column=1,value="Summary").font = Font(bold=True, size=11)
    valid = res_df.dropna(subset=['direct_slope'])
    for i, (l, v) in enumerate([
        ("National multiplier", f"{NAT_MULT:.3f}"),
        ("DC industry share", f"{100/NAT_MULT:.2f}%"),
        ("Median local_share", f"{res_df['local_share'].median():.1%}"),
        ("",""),
        ("Valid regressions", f"{len(valid)} / {len(res_df)}"),
        (f"Median β cleaned direct (jobs/{x_unit})", f"{valid['direct_slope'].median():,.1f}"),
        (f"Median β cleaned×mult (jobs/{x_unit})", f"{valid['mult_slope'].median():,.1f}"),
        ("p<0.05 (cleaned direct)", f"{(valid['direct_p_value']<0.05).sum()} / {len(valid)}"),
        ("p<0.05 (cleaned×mult)", f"{(valid['mult_p_value']<0.05).sum()} / {len(valid)}"),
        ("",""),
        ("Method", "cleaned_direct = NAICS_5182 × local_share"),
        ("", "cleaned_multiplied = cleaned_direct × local_multiplier"),
        ("local_share", "PwC_direct / (PwC_direct + spillover/nat_mult)"),
    ]):
        ws.cell(row=sr+1+i,column=1,value=l).font = Font(bold=True if l else False, size=10)
        ws.cell(row=sr+1+i,column=3,value=v)

    for col, w in {'A':7,'B':22,'C':5,'D':8,'E':8,'F':9,'G':9,
                   'H':13,'I':10,'J':13,'K':13,'L':12,'M':8,'N':9,
                   'O':13,'P':10,'Q':13,'R':13,'S':12,'T':8,'U':9}.items():
        ws.column_dimensions[col].width = w


def write_panel_sheet(wb, panel_df):
    ws = wb.create_sheet("Panel Data")
    ws.sheet_properties.tabColor = "70AD47"
    ws.freeze_panes = "A2"
    hf = Font(bold=True, color="FFFFFF", size=11)

    headers = ['State', 'State Name', 'Year', 'NAICS 5182\n(observed)',
               'DC Count', 'DC Cap\n(GW)', 'Local\nShare', 'Local\nMultiplier',
               'Cleaned\nDirect', 'Cleaned\n×Multiplier', 'Raw\nMultiplied']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = PatternFill("solid", fgColor="70AD47")
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 35

    fmts = [None, None, None, '#,##0.0', '#,##0', '0.000', '0.0%', '0.00', '#,##0.0', '#,##0', '#,##0']
    for idx, (_, r) in enumerate(panel_df.sort_values(['state_abbr','year']).iterrows()):
        rn = idx + 2
        vals = [r['state_abbr'], r['state_name'], int(r['year']), r['emp_5182'],
                int(r['dc_count']), r['dc_capacity_GW'], r['local_share'], r['local_multiplier'],
                r['emp_cleaned_direct'], r['emp_cleaned_multiplied'], r['emp_raw_multiplied']]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=rn, column=ci, value=v)
            c.border = BORDER
            if fmt: c.number_format = fmt

    for ci, w in enumerate([7,22,7,12,9,10,8,8,12,12,12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# Build workbook
wb = Workbook()
wb.remove(wb.active)

write_raw_sheet(wb, "Raw OLS by Capacity", "4472C4", res_raw_cap,
                "DC Capacity", "GW", "0.000", "4472C4", "2F5496", "D6E4F0", "D6E4F0", "#,##0")
write_panel_sheet(wb, panel)
write_cleaned_sheet(wb, "Cleaned OLS by Capacity", "7030A0", res_cln_cap,
                    "DC Capacity", "GW", "0.000", "7030A0", "5B259F", "#,##0")

# ============================================================
# 6. Panel Fixed Effects: Emp_it = α_i + γ_t + β × X_it + ε_it
# ============================================================
def run_panel_fe(df, x_col, y_col, entity_col='state_abbr', time_col='year'):
    pdf = df[[entity_col, time_col, x_col, y_col]].dropna().copy()
    pdf = pdf.set_index([entity_col, time_col])
    mod = PanelOLS(pdf[y_col], pdf[[x_col]], entity_effects=True, time_effects=True)
    return mod.fit(cov_type='clustered', cluster_entity=True)


fe_specs = []
for y_col, y_label in [
    ('emp_5182', 'Raw NAICS 5182'),
    ('emp_cleaned_direct', 'Cleaned Direct'),
    ('emp_raw_multiplied', 'Raw × Multiplier+Spillover'),
    ('emp_cleaned_multiplied', 'Cleaned × Local Multiplier'),
]:
    res = run_panel_fe(panel, 'dc_capacity_GW', y_col)
    fe_specs.append({'x_col': 'dc_capacity_GW', 'x_label': 'Capacity (GW)',
                     'y_col': y_col, 'y_label': y_label, 'result': res})


def write_panel_fe_sheet(wb):
    ws = wb.create_sheet("Panel FE")
    ws.sheet_properties.tabColor = "1F4E79"
    ws.freeze_panes = "A3"
    hf = Font(bold=True, color="FFFFFF", size=11)

    ws.merge_cells('A1:I1')
    c = ws['A1']
    c.value = "Employment — Panel Fixed Effects: Emp_it = α_i + γ_t + β × X_it + ε_it"
    c.font = hf; c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal='center')

    headers = ['Specification (Y)', 'X Variable', 'β', 'SE\n(clustered)', 'CI95\nLower',
               'CI95\nUpper', 'p-value', 'R² within', 'N obs / States']
    sub_fill = PatternFill("solid", fgColor="D6E4F0")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER; cell.fill = sub_fill
    ws.row_dimensions[2].height = 35

    for idx, sp in enumerate(fe_specs):
        rn = idx + 3
        res_fe = sp['result']
        x = sp['x_col']
        beta = res_fe.params[x]
        se = res_fe.std_errors[x]
        pval = res_fe.pvalues[x]

        vals = [sp['y_label'], sp['x_label'],
                beta, se, beta - 1.96 * se, beta + 1.96 * se,
                pval, res_fe.rsquared_within,
                f"{res_fe.nobs} / {res_fe.entity_info['total']}"]
        fmts = [None, None, '#,##0.0', '#,##0.0', '#,##0.0', '#,##0.0', '0.0000', '0.000', None]

        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=rn, column=ci, value=v)
            cell.border = BORDER
            if fmt:
                cell.number_format = fmt
        if pval < 0.05:
            for ci in [3, 4, 5, 6, 7]:
                ws.cell(row=rn, column=ci).fill = SIG_FILL

    sr = len(fe_specs) + 4
    ws.cell(row=sr, column=1, value="Notes").font = Font(bold=True, size=11)
    for i, (l, v) in enumerate([
        ("Model", "Emp_it = α_i + γ_t + β × X_it + ε_it"),
        ("Fixed Effects", "State (entity) + Year (time)"),
        ("SE", "Clustered by state"),
        ("Period", "2016–2024 (9 years)"),
        ("Interpretation", "β = avg within-state employment change per unit X,"),
        ("", "controlling for state-level and time-level unobservables"),
    ]):
        ws.cell(row=sr + 1 + i, column=1, value=l).font = Font(bold=True if l else False, size=10)
        ws.cell(row=sr + 1 + i, column=3, value=v)

    for ci, w in enumerate([28, 16, 12, 12, 12, 12, 10, 10, 14], 1):
        ws.column_dimensions[chr(64 + ci)].width = w


write_panel_fe_sheet(wb)

wb.save(_DIR / '02_ols_employment_dc_capacity.xlsx')
print("Saved employment/02_ols_employment_dc_capacity.xlsx")

for label, res in [("Raw by Capacity", res_raw_cap),
                   ("Cleaned by Capacity", res_cln_cap)]:
    v = res.dropna(subset=['direct_slope'])
    print(f"\n=== {label} ===")
    print(f"  Valid: {len(v)}/{len(res)}")
    print(f"  Median β direct: {v['direct_slope'].median():,.1f}")
    print(f"  Median β multiplied: {v['mult_slope'].median():,.1f}")
    print(f"  p<0.05: {(v['direct_p_value']<0.05).sum()}/{len(v)} direct, "
          f"{(v['mult_p_value']<0.05).sum()}/{len(v)} mult")

print("\n=== Panel FE ===")
for sp in fe_specs:
    res_fe = sp['result']
    x = sp['x_col']
    beta = res_fe.params[x]
    p = res_fe.pvalues[x]
    sig = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else ''
    print(f"  {sp['y_label']:35s} ~ {sp['x_label']:15s}  β={beta:>10,.1f}  p={p:.4f}{sig}")
