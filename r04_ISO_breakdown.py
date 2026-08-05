import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

# ================= Configuration =================
EIA_DIR         = Path("./rider/EIA")
PATH_PRICES_OLD = EIA_DIR / "EIA_AEO2023_prices_by_service.xlsx"
PATH_PRICES_NEW = EIA_DIR / "EIA_AEO2025_prices_by_service.xlsx"
PATH_MAP        = EIA_DIR / "EIA_area.xlsx"
MAP_SHEET       = 0
OUT_DIR         = EIA_DIR
YEAR_CUTOFF     = 2024

# --- DC price impact ---
PATH_DC_IMPACT     = Path("./results/r3_summary/table3_zone_impact_2025.csv")
PATH_DC_CAP_ISO    = Path("./tables/dc_cumulative_by_iso.xlsx")
PATH_DC_CAP_CITY   = Path("./tables_city/dc_cumulative_by_city.xlsx")
DC_YEARS           = list(range(2020, 2031))
DC_BASE_YEAR       = 2019
DC_SCALE           = 0.1        # $/MWh -> cents/kWh
DC_ROW_NAME        = "DC_cumulate"
TOTAL_ROW_NAME     = "total"
TOTAL_MINUS_ROW    = "total_minus_DC"

REGION_TO_ISO = {
    "CAISO": "CAISO",
    "ERCOT": "ERCOT",
    "MISO":  "MISO",
    "PJM":   "PJM",
    "Non-ISO Cities": "city",
}
ZONE_NAME_OVERRIDE = {
    "EKPC PJM": "EKPC",
}
# Map zone column names in dc_cumulative to table3 zone names (tolerates spacing and punctuation differences)
CAP_COL_OVERRIDE = {
    "EKPC PJM": "EKPC",
}
# ===================================================

ROW_KEYS = {
    "gen": ["gen", "generation"],
    "tre": ["tre", "transmission"],
    "dse": ["dse", "distribution"],
}


def _norm(s: str) -> str:
    if pd.isna(s):
        return ""
    s = str(s)
    s = re.sub(r"\(.*?\)", "", s)
    return s.strip().lower()


def _which_row(name: str) -> str | None:
    n = _norm(name)
    for k, kws in ROW_KEYS.items():
        for kw in kws:
            if n.startswith(kw) or kw in n:
                return k
    return None


def read_price_sheet(area_name: str, df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["area", "kind", "year", "value"])
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    df.rename(columns={df.columns[0]: "item"}, inplace=True)
    df["kind"] = df["item"].apply(_which_row)
    df = df.loc[df["kind"].notna()].drop(columns=["item"]).reset_index(drop=True)

    year_cols = [c for c in df.columns if c != "kind"]
    keep_cols = []
    for c in year_cols:
        cnum = re.sub(r"[^\d]", "", str(c))
        if len(cnum) >= 4 and cnum[:4].isdigit():
            keep_cols.append(c)
    if not keep_cols:
        keep_cols = [c for c in df.columns if c != "kind"]

    df = df[["kind"] + keep_cols]
    out = df.melt(id_vars="kind", var_name="year", value_name="value")
    out["year"] = out["year"].astype(str).str.extract(r"(\d{4})").squeeze()
    out = out.dropna(subset=["year"]).copy()
    out["area"] = area_name
    out["value"] = pd.to_numeric(out["value"], errors="coerce")
    return out[["area", "kind", "year", "value"]]


def read_price_workbook(path: Path) -> pd.DataFrame:
    xl = pd.ExcelFile(path)
    frames = []
    for area in xl.sheet_names:
        frames.append(read_price_sheet(area, xl.parse(area, header=0)))
    long = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=["area", "kind", "year", "value"])
    return long


def merge_prices_with_updates(old_long: pd.DataFrame, new_long: pd.DataFrame, cutoff_year: int) -> pd.DataFrame:
    old_long = old_long.copy()
    new_long = new_long.copy()
    old_long["year"] = old_long["year"].astype(int)
    new_long["year"] = new_long["year"].astype(int)

    old_pre = old_long.loc[old_long["year"] < cutoff_year]
    new_post = new_long.loc[new_long["year"] >= cutoff_year]

    combined = pd.concat([old_pre, new_post], ignore_index=True)
    combined = (combined
                .sort_values(["area", "kind", "year"])
                .drop_duplicates(subset=["area", "kind", "year"], keep="last")
                .reset_index(drop=True))
    return combined


def pivot_area_year(combined_long: pd.DataFrame) -> pd.DataFrame:
    piv = (combined_long
           .pivot_table(index=["area", "year"], columns="kind", values="value", aggfunc="mean")
           .reset_index())
    for col in ["gen", "tre", "dse"]:
        if col not in piv.columns:
            piv[col] = pd.NA
    return piv


def try_col(cands, df):
    for c in cands:
        if c in df.columns:
            return c
    return None


def _normalize_zone_key(z: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(z).strip()).upper()


def _match_cap_col(cap_df_cols: list[str], zone_name: str) -> str | None:
    """Find the dc_cumulative column matching zone_name."""
    zone_norm = _normalize_zone_key(zone_name)
    for c in cap_df_cols:
        if _normalize_zone_key(c) == zone_norm:
            return c
    return None


def load_dc_impact_timeseries(
    path_table3: Path,
    path_cap_iso: Path,
    path_cap_city: Path,
) -> dict[str, dict[str, dict[int, float]]]:
    """
    Returns {iso_file_key: {sheet_zone_norm: {year: dP_cents_kwh}}}
    dP_year = beta * max(0, (cap_year - cap_base)/1000 - DC_start_GW) * DC_SCALE
    """
    t3 = pd.read_csv(path_table3)
    cap_iso = pd.read_excel(path_cap_iso, sheet_name=None)
    cap_city = pd.read_excel(path_cap_city, sheet_name=0).set_index("Year")

    result: dict[str, dict[str, dict[int, float]]] = {}

    for _, row in t3.iterrows():
        region = str(row["Region"]).strip()
        iso_key = REGION_TO_ISO.get(region)
        if iso_key is None:
            continue

        zone_raw = str(row["Zone"]).strip()
        zone_sheet = ZONE_NAME_OVERRIDE.get(zone_raw, zone_raw)
        zone_norm = _normalize_zone_key(zone_sheet)
        beta = float(row["beta"])
        dc_start = float(row["DC_start_GW"])

        if region == "Non-ISO Cities":
            cap_df = cap_city
        else:
            cap_df = cap_iso.get(region)
            if cap_df is None:
                continue
            cap_df = cap_df.set_index("Year")

        cap_col = _match_cap_col(list(cap_df.columns), zone_raw)
        if cap_col is None:
            cap_col = _match_cap_col(list(cap_df.columns), zone_sheet)

        yearly = {}
        for y in DC_YEARS:
            if cap_col is not None and y in cap_df.index and DC_BASE_YEAR in cap_df.index:
                cap_y = float(cap_df.loc[y, cap_col])
                cap_base = float(cap_df.loc[DC_BASE_YEAR, cap_col])
                dDC = (cap_y - cap_base) / 1000.0 - dc_start
            else:
                dDC = 0.0
            dp = beta * max(0.0, dDC) * DC_SCALE
            yearly[y] = round(dp, 6)

        result.setdefault(iso_key, {})[zone_norm] = yearly

    return result


def add_dc_rows_to_xlsx(xlsx_path: Path, zone_impacts: dict[str, dict[int, float]]):
    """Append DC_cumulate and total_minus_DC rows to each matching sheet of an existing xlsx."""
    wb = load_workbook(xlsx_path)
    sheet_norm_map = {_normalize_zone_key(ws.title): ws.title for ws in wb.worksheets}

    for zone_norm, yearly in zone_impacts.items():
        ws_title = sheet_norm_map.get(zone_norm)
        if ws_title is None:
            print(f"  [warning] {xlsx_path.name}: no sheet matches zone '{zone_norm}'")
            continue
        ws = wb[ws_title]

        # Build the year to column map
        year_col_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            try:
                year_col_map[int(str(v).strip())] = c
            except Exception:
                pass
        next_col = ws.max_column + 1
        for y in sorted(yearly.keys()):
            if y not in year_col_map:
                ws.cell(1, next_col).value = y
                year_col_map[y] = next_col
                next_col += 1

        def find_row(name):
            for r in range(2, ws.max_row + 2):
                v = ws.cell(r, 1).value
                if isinstance(v, str) and v.strip().lower() == name.strip().lower():
                    return r
            return None

        def find_or_create_row(name):
            r = find_row(name)
            if r is not None:
                return r
            new_r = ws.max_row + 1
            ws.cell(new_r, 1).value = name
            return new_r

        dc_row = find_or_create_row(DC_ROW_NAME)
        total_row = find_row(TOTAL_ROW_NAME)
        minus_row = find_or_create_row(TOTAL_MINUS_ROW)

        for y, dp_value in yearly.items():
            c = year_col_map[y]
            ws.cell(dc_row, c).value = dp_value

            col_letter = get_column_letter(c)
            if total_row is not None:
                ws.cell(minus_row, c).value = f"={col_letter}{total_row}-{col_letter}{dc_row}"
            else:
                ws.cell(minus_row, c).value = -dp_value

    wb.save(xlsx_path)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Read and merge the old and new prices
    old_long = read_price_workbook(PATH_PRICES_OLD)
    new_long = read_price_workbook(PATH_PRICES_NEW)
    combined_long = merge_prices_with_updates(old_long, new_long, YEAR_CUTOFF)
    price_piv = pivot_area_year(combined_long)

    # Map market area to ISO and zones
    map_raw = pd.read_excel(PATH_MAP, sheet_name=MAP_SHEET, header=0)

    COL_AREA = try_col(["Abbr", "Market Area", "Area", "Abbr 1", "Abbr1", "MA"], map_raw) or map_raw.columns[0]
    COL_ISO  = try_col(["ISO", "Iso"], map_raw) or map_raw.columns[3]
    COL_ZONE = try_col(["Zone", "Zones", "Load Zone", "Load Zones"], map_raw) or map_raw.columns[4]

    mapping = (map_raw[[COL_AREA, COL_ISO, COL_ZONE]]
               .rename(columns={COL_AREA: "area", COL_ISO: "ISO", COL_ZONE: "zones"})
               .dropna(subset=["area", "ISO", "zones"])
               .assign(zone_list=lambda d: d["zones"].astype(str).str.split(","))
               .explode("zone_list")
               .assign(zone=lambda d: d["zone_list"].str.strip())
               .drop(columns=["zones", "zone_list"])
               .dropna(subset=["zone"])
               .reset_index(drop=True))

    # Merge and average within ISO-zone-year
    merged = mapping.merge(price_piv, on="area", how="left")
    for col in ["gen", "tre", "dse"]:
        merged[col] = pd.to_numeric(merged[col], errors="coerce")

    avg_by_zone = (merged
                   .groupby(["ISO", "zone", "year"], as_index=False)[["gen", "tre", "dse"]]
                   .mean())

    avg_by_zone["total"] = avg_by_zone[["gen", "tre", "dse"]].sum(axis=1, min_count=1)
    avg_by_zone["gen_share"] = avg_by_zone["gen"] / avg_by_zone["total"]

    # Write the baseline price table by ISO
    for iso, df_iso in avg_by_zone.groupby("ISO"):
        out_path = OUT_DIR / f"{iso}_zone_prices.xlsx"
        with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
            for zone, df_zone in df_iso.groupby("zone"):
                years = sorted({str(int(y)) for y in df_zone["year"].dropna().astype(int).unique()})
                rows = []
                for metric in ["gen", "tre", "dse", "total", "gen_share"]:
                    row = [metric]
                    for y in years:
                        v = df_zone.loc[df_zone["year"].astype(int) == int(y), metric]
                        row.append(v.values[0] if len(v) else pd.NA)
                    rows.append(row)
                out_df = pd.DataFrame(rows, columns=["item"] + years)
                name_map = {
                    "gen": "gen (Generation)",
                    "tre": "tre (Transmission)",
                    "dse": "dse (Distribution)",
                    "total": "total",
                    "gen_share": "gen_share (=gen/total)",
                }
                out_df["item"] = out_df["item"].map(name_map).fillna(out_df["item"])
                safe_sheet = re.sub(r"[/\\?*\[\]:]", "_", str(zone))[:31] or "zone"
                out_df.to_excel(writer, index=False, sheet_name=safe_sheet)
        print(f"wrote {iso} -> {out_path}")

    # Append DC impact rows to each ISO xlsx (2020-2030)
    dc_impacts = load_dc_impact_timeseries(PATH_DC_IMPACT, PATH_DC_CAP_ISO, PATH_DC_CAP_CITY)
    for iso_key, zone_map in dc_impacts.items():
        xlsx_path = OUT_DIR / f"{iso_key}_zone_prices.xlsx"
        if not xlsx_path.exists():
            print(f"[skip] {xlsx_path} does not exist")
            continue
        add_dc_rows_to_xlsx(xlsx_path, zone_map)
        print(f"DC impact write {xlsx_path.name}: {len(zone_map)} zones, years {DC_YEARS[0]}-{DC_YEARS[-1]}")


if __name__ == "__main__":
    main()
