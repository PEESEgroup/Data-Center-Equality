"""
Construction & related industry employment regressions against DC capacity.
State-level OLS (2016-2024) + Panel Fixed Effects.

NAICS codes:
  23:   Construction (total)
  2362: Nonresidential Building Construction
  5415: Computer Systems Design & Related Services

Output: 04_construction_analysis.xlsx
  Sheet 1-3: OLS for each NAICS (by capacity)
  Sheet 4: Cross-NAICS Comparison (top states)
  Sheet 5: Panel FE
"""

import pandas as pd
import numpy as np
from pathlib import Path
from scipy import stats
from linearmodels.panel import PanelOLS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 1. Load data
# ============================================================
_DIR = Path(__file__).resolve().parent / "employment"

qwi = pd.read_csv(_DIR / 'qwi_all_naics_annual.csv', dtype={'naics': str})
dc = pd.read_csv(_DIR / 'dc_facilities_by_state_year.csv')

STATE_ABBR_MAP = {
    'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA',
    'Colorado':'CO','Connecticut':'CT','Delaware':'DE','District of Columbia':'DC',
    'Florida':'FL','Georgia':'GA','Hawaii':'HI','Idaho':'ID','Illinois':'IL',
    'Indiana':'IN','Iowa':'IA','Kansas':'KS','Kentucky':'KY','Louisiana':'LA',
    'Maine':'ME','Maryland':'MD','Massachusetts':'MA','Michigan':'MI','Minnesota':'MN',
    'Mississippi':'MS','Missouri':'MO','Montana':'MT','Nebraska':'NE','Nevada':'NV',
    'New Hampshire':'NH','New Jersey':'NJ','New Mexico':'NM','New York':'NY',
    'North Carolina':'NC','North Dakota':'ND','Ohio':'OH','Oklahoma':'OK','Oregon':'OR',
    'Pennsylvania':'PA','Rhode Island':'RI','South Carolina':'SC','South Dakota':'SD',
    'Tennessee':'TN','Texas':'TX','Utah':'UT','Vermont':'VT','Virginia':'VA',
    'Washington':'WA','West Virginia':'WV','Wisconsin':'WI','Wyoming':'WY'
}
ABBR_TO_NAME = {v: k for k, v in STATE_ABBR_MAP.items()}
YEARS = list(range(2016, 2025))

# ============================================================
# 2. Build DC panel
# ============================================================
dc_states = dc[dc['state_abbr'] != 'US'].copy()
dc_panel = []
for _, row in dc_states.iterrows():
    for yr in YEARS:
        dc_panel.append({
            'state_abbr': row['state_abbr'], 'year': yr,
            'dc_count': row[f'count_{yr}'],
            'dc_capacity_GW': row[f'MW_{yr}'] / 1000
        })
dc_panel = pd.DataFrame(dc_panel)

# ============================================================
# 3. Build employment panels
# ============================================================
qwi['state_abbr'] = qwi['state_name'].map(STATE_ABBR_MAP)

NAICS_SPECS = [
    ('23', 'Construction (Total)', 'NAICS 23 — Construction (Total)'),
    ('2362', 'Nonresidential Building Construction', 'NAICS 2362 — Nonresidential Building Construction'),
    ('5415', 'Computer Systems Design', 'NAICS 5415 — Computer Systems Design'),
]

panels = {}
for naics, _, _ in NAICS_SPECS:
    q = qwi[qwi['naics'] == naics][['state_abbr', 'year', 'Emp_annual_avg']].copy()
    p = q.merge(dc_panel, on=['state_abbr', 'year'], how='inner')
    p = p.rename(columns={'Emp_annual_avg': f'emp_{naics}'})
    panels[naics] = p

# ============================================================
# 4. State-level OLS
# ============================================================
def run_state_ols(panel_df, x_col, y_col):
    results = []
    for abbr in sorted(panel_df['state_abbr'].unique()):
        df_st = panel_df[panel_df['state_abbr'] == abbr].dropna(subset=[y_col, x_col])
        n = len(df_st)
        x = df_st[x_col].values.astype(float)
        y = df_st[y_col].values.astype(float)

        base = {
            'state_abbr': abbr,
            'state_name': ABBR_TO_NAME.get(abbr, abbr),
            'n_obs': n,
            'x_min': x.min() if len(x) > 0 else np.nan,
            'x_max': x.max() if len(x) > 0 else np.nan,
        }

        if n < 3 or np.std(x) < 1e-10:
            for f in ['slope', 'se', 'ci_low', 'ci_high', 'r_squared', 'p_value']:
                base[f] = np.nan
            results.append(base)
            continue

        sl, ic, r, p, se = stats.linregress(x, y)
        ci = 1.96 * se
        base.update({
            'slope': sl, 'se': se, 'ci_low': sl - ci, 'ci_high': sl + ci,
            'r_squared': r ** 2, 'p_value': p,
        })
        results.append(base)
    return pd.DataFrame(results)


# ============================================================
# 5. Panel FE
# ============================================================
def run_panel_fe(df, x_col, y_col):
    pdf = df[['state_abbr', 'year', x_col, y_col]].dropna().copy()
    pdf = pdf.set_index(['state_abbr', 'year'])
    mod = PanelOLS(pdf[y_col], pdf[[x_col]], entity_effects=True, time_effects=True)
    return mod.fit(cov_type='clustered', cluster_entity=True)


# ============================================================
# 6. Run all regressions
# ============================================================
ols_results = {}
fe_results = {}

for naics, _, _ in NAICS_SPECS:
    p = panels[naics]
    y_col = f'emp_{naics}'
    ols_results[naics] = run_state_ols(p, 'dc_capacity_GW', y_col)
    fe_results[naics] = run_panel_fe(p, 'dc_capacity_GW', y_col)

# ============================================================
# 7. Write Excel
# ============================================================
thin = Side(style='thin', color='000000')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
SIG_FILL = PatternFill("solid", fgColor="C6EFCE")
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")

wb = Workbook()
wb.remove(wb.active)

# --- Sheets 1-3: OLS by NAICS ---
colors = [('4472C4', 'D6E4F0'), ('548235', 'E2EFDA'), ('7030A0', 'E8D5F5')]

for (naics, short_label, full_title), (hdr_color, fill_color) in zip(NAICS_SPECS, colors):
    ws = wb.create_sheet(f"OLS NAICS {naics}")
    ws.sheet_properties.tabColor = hdr_color
    ws.freeze_panes = "A3"
    hf = Font(bold=True, color="FFFFFF", size=11)

    ws.merge_cells('A1:B1')
    c = ws['A1']; c.value = full_title; c.font = hf
    c.fill = PatternFill("solid", fgColor=hdr_color); c.alignment = Alignment(horizontal='center')
    ws.merge_cells('C1:I1')
    c = ws['C1']; c.value = 'OLS by DC Capacity (GW)'; c.font = hf
    c.fill = PatternFill("solid", fgColor=hdr_color); c.alignment = Alignment(horizontal='center')

    headers = ['State', 'State Name',
               'β (jobs/GW)', 'SE', 'CI95 Lower', 'CI95 Upper', 'R²', 'p-value', 'Cap Range\n(GW)']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        cell.fill = PatternFill("solid", fgColor=fill_color)
    ws.row_dimensions[2].height = 35

    res_cap = ols_results[naics].sort_values('slope', ascending=False, na_position='last')

    for idx, (_, r) in enumerate(res_cap.iterrows()):
        rn = idx + 3
        ws.cell(row=rn, column=1, value=r['state_abbr']).border = BORDER
        ws.cell(row=rn, column=2, value=r['state_name']).border = BORDER

        for ci, (f, fmt) in enumerate(zip(
            ['slope', 'se', 'ci_low', 'ci_high', 'r_squared', 'p_value'],
            ['#,##0', '#,##0', '#,##0', '#,##0', '0.000', '0.0000']
        ), 3):
            v = r[f]
            cell = ws.cell(row=rn, column=ci, value=v if not pd.isna(v) else None)
            cell.number_format = fmt; cell.border = BORDER

        cap_range = f"{r['x_min']:.3f}–{r['x_max']:.3f}" if not pd.isna(r['x_min']) else ''
        ws.cell(row=rn, column=9, value=cap_range).border = BORDER

        if not pd.isna(r['p_value']) and r['p_value'] < 0.05:
            fl = SIG_FILL if r['slope'] > 0 else NEG_FILL
            for ci in [3, 4, 8]: ws.cell(row=rn, column=ci).fill = fl

    for ci, w in enumerate([5, 20, 11, 9, 11, 11, 7, 9, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

# --- Sheet 4: Cross-NAICS Comparison ---
ws4 = wb.create_sheet("Cross-NAICS Comparison")
ws4.sheet_properties.tabColor = "C00000"
ws4.freeze_panes = "A3"
hf = Font(bold=True, color="FFFFFF", size=11)

groups4 = [
    ('A1:D1', 'State Info', '4472C4'),
    ('E1:G1', 'NAICS 23\nConstruction (Total)', '4472C4'),
    ('H1:J1', 'NAICS 2362\nNonresidential Building Construction', '548235'),
    ('K1:M1', 'NAICS 5415\nComputer Systems Design', '7030A0'),
]
for rng, title, color in groups4:
    ws4.merge_cells(rng)
    c = ws4[rng.split(':')[0]]
    c.value = title; c.font = hf
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

headers4 = ['State', 'State Name', 'DC Cap\n2024 (GW)', 'DC\nCount 2024',
            'β/GW', 'p-value', 'R²', 'β/GW', 'p-value', 'R²', 'β/GW', 'p-value', 'R²']
for ci, h in enumerate(headers4, 1):
    cell = ws4.cell(row=2, column=ci, value=h)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER

# Get 2024 capacity/count
dc_latest = dc_panel[dc_panel['year'] == 2024][['state_abbr', 'dc_capacity_GW', 'dc_count']]
dc_latest_dict = dict(zip(dc_latest['state_abbr'], dc_latest[['dc_capacity_GW', 'dc_count']].to_dict('records')))

# Sort by capacity
all_states = sorted(ols_results['23']['state_abbr'].unique())
state_caps = [(s, dc_latest_dict.get(s, {}).get('dc_capacity_GW', 0)) for s in all_states]
state_caps.sort(key=lambda x: -x[1])

for idx, (abbr, cap) in enumerate(state_caps[:15]):
    rn = idx + 3
    ws4.cell(row=rn, column=1, value=abbr).border = BORDER
    ws4.cell(row=rn, column=2, value=ABBR_TO_NAME.get(abbr, '')).border = BORDER
    c = ws4.cell(row=rn, column=3, value=cap); c.number_format = '0.000'; c.border = BORDER
    cnt = dc_latest_dict.get(abbr, {}).get('dc_count', None)
    ws4.cell(row=rn, column=4, value=cnt).border = BORDER

    for ni, naics in enumerate(['23', '2362', '5415']):
        res = ols_results[naics]
        r = res[res['state_abbr'] == abbr]
        if len(r) > 0:
            r = r.iloc[0]
            for ci, (f, fmt) in enumerate(zip(['slope', 'p_value', 'r_squared'],
                                               ['#,##0', '0.000', '0.000']), 5 + ni * 3):
                v = r[f]
                cell = ws4.cell(row=rn, column=ci, value=v if not pd.isna(v) else None)
                cell.number_format = fmt; cell.border = BORDER

for ci, w in enumerate([5, 16, 10, 8, 10, 8, 7, 10, 8, 7, 10, 8, 7], 1):
    ws4.column_dimensions[get_column_letter(ci)].width = w

# --- Sheet 5: Panel FE ---
ws5 = wb.create_sheet("Panel FE")
ws5.sheet_properties.tabColor = "1F4E79"
ws5.freeze_panes = "A3"

ws5.merge_cells('A1:I1')
c = ws5['A1']
c.value = "Construction & Related — Panel Fixed Effects: Emp_it = α_i + γ_t + β × X_it + ε_it"
c.font = Font(bold=True, color="FFFFFF", size=11)
c.fill = PatternFill("solid", fgColor="1F4E79")
c.alignment = Alignment(horizontal='center')

headers5 = ['Specification (Y)', 'X Variable', 'β', 'SE\n(clustered)', 'CI95\nLower',
            'CI95\nUpper', 'p-value', 'R² within', 'N obs / States']
for ci, h in enumerate(headers5, 1):
    cell = ws5.cell(row=2, column=ci, value=h)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
    cell.fill = PatternFill("solid", fgColor="D6E4F0")
ws5.row_dimensions[2].height = 35

for fe_idx, (naics, short_label, _) in enumerate(NAICS_SPECS):
    rn = fe_idx + 3
    res = fe_results[naics]
    beta = res.params['dc_capacity_GW']
    se = res.std_errors['dc_capacity_GW']
    pval = res.pvalues['dc_capacity_GW']

    vals = [f'{short_label} (NAICS {naics})', 'Capacity (GW)',
            beta, se, beta - 1.96 * se, beta + 1.96 * se,
            pval, res.rsquared_within,
            f"{res.nobs} / {res.entity_info['total']}"]
    fmts5 = [None, None, '#,##0.0', '#,##0.0', '#,##0.0', '#,##0.0', '0.0000', '0.000', None]

    for ci, (v, fmt) in enumerate(zip(vals, fmts5), 1):
        cell = ws5.cell(row=rn, column=ci, value=v)
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt

    if pval < 0.05:
        for ci in [3, 4, 5, 6, 7]:
            ws5.cell(row=rn, column=ci).fill = SIG_FILL

# Notes
sr = len(NAICS_SPECS) + 4
ws5.cell(row=sr, column=1, value="Notes").font = Font(bold=True, size=11)
for i, (l, v) in enumerate([
    ("Model", "Emp_it = α_i + γ_t + β × X_it + ε_it"),
    ("Fixed Effects", "State (entity) + Year (time)"),
    ("SE", "Clustered by state"),
    ("Period", "2016–2024 (9 years)"),
]):
    ws5.cell(row=sr + 1 + i, column=1, value=l).font = Font(bold=True, size=10)
    ws5.cell(row=sr + 1 + i, column=3, value=v)

for ci, w in enumerate([28, 16, 12, 12, 12, 12, 10, 10, 14], 1):
    ws5.column_dimensions[chr(64 + ci)].width = w

wb.save(_DIR / '04_construction_analysis.xlsx')
print("Saved 04_construction_analysis.xlsx")

# Print summary
for naics, short_label, _ in NAICS_SPECS:
    res = fe_results[naics]
    beta = res.params['dc_capacity_GW']
    p = res.pvalues['dc_capacity_GW']
    sig = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else ''
    print(f"  FE {short_label:25s} ~ Capacity   β={beta:>10,.1f}  p={p:.4f}{sig}")
