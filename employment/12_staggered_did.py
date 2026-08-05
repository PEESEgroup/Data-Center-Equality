"""
Staggered DiD (Callaway & Sant'Anna 2021)
Treatment: first year NEW DC capacity (added since 2016) exceeds threshold.
Control: never-treated states.

Thresholds: 500MW, 1GW
Y variables: employment (raw 5182, cleaned direct, cleaned × multiplier)
             + construction (NAICS 23, 2362, 5415)

Output: 06_staggered_did_analysis.xlsx
  Sheet 1: Overall ATT
  Sheet 2: Event Study (dynamic ATTs by relative time)
  Sheet 3: Treatment Cohorts
  Sheet 4: Event Study Plot Data (wide format for charting)

IMPLEMENTATION NOTES
---------------------------------------------------------
1. UNIVERSE (C3, M6).  The universe is imported from
   `employment/analysis_universe.py`: the 48 contiguous states plus DC, 49
   units.  The published run attached a 51-unit treatment assignment
   (`get_treatment_years` iterated the capacity file's 51 units, giving
   18 + 33 and 13 + 38) to a 50-unit, DC-less outcome panel, because
   STATE_ABBR_MAP mapped the literal string "DC" to NaN.  Treatment assignment
   and outcome panel now share one universe by construction, and it is checked.

2. THE aggte ALIASING BUG.
   `ATTgt.aggte()` returns `self`, so `agg_s = model.aggte('simple')` and
   `agg_d = model.aggte('dynamic')` are THE SAME OBJECT.  The published code
   called both and only then read `agg_s.atte['overall_att']`, by which point
   the simple aggregation had been overwritten by the dynamic one.  Every
   number in the published "Overall ATT" sheet is therefore the DYNAMIC
   (event-study) overall ATT, mislabelled as the simple aggregation.
   Verified: on the published sample the simple ATT for 500 MW / raw NAICS 5182
   is 1,994.23 and the dynamic overall is 1,509.30; the published cell reads
   1,509.30.  Both aggregations are now read BEFORE the other is requested and
   both are written out, in labelled columns.  `ATT` is kept as the dynamic
   overall so the published series remains traceable; `ATT_simple` is new.

3. REPRODUCIBILITY.  The csdid standard errors are bootstrapped and the
   published run set no seed, so the published SEs cannot be reproduced exactly
   -- only to within Monte Carlo error.  A seed is set here.  The ATT point
   estimates ARE deterministic and reproduce exactly, but only with the pinned
   csdid==0.2.9 of requirements.txt: csdid 0.4.2 changes them by up to 6.5 per
   cent (500 MW raw 5182 1,509.30 -> 1,512.81; 1 GW raw 5182 3,824.72 ->
   4,074.72).  The installed version is asserted at import.
"""

import pandas as pd
import numpy as np
import warnings
import io
import sys
from pathlib import Path
from collections import Counter
from scipy.stats import norm
from csdid import att_gt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

BOOTSTRAP_SEED = 20260728
REQUIRED_CSDID = "0.2.9"          # requirements.txt pin; see docstring item 3

import importlib.metadata as _md
_csdid_version = _md.version("csdid")
if _csdid_version != REQUIRED_CSDID:
    raise SystemExit(
        f"csdid {_csdid_version} is installed but this analysis is pinned to "
        f"csdid=={REQUIRED_CSDID} (requirements.txt).  The ATT point "
        f"estimates are NOT stable across csdid versions: 0.4.2 moves the 1 GW "
        f"raw NAICS 5182 ATT from 3,824.72 to 4,074.72, a 6.5 per cent change. "
        f"Run  python3 -m pip install csdid=={REQUIRED_CSDID}  first.")

# Release-set path rule: no absolute paths, and no assumption about the
# working directory.  This script now lives IN employment.
_DIR = Path(__file__).resolve().parent                    # employment
sys.path.insert(0, str(_DIR))
from analysis_universe import (N_ANALYSIS_UNITS, OUT_OF_SCOPE_UNITS,
                               UNIVERSE_LABEL, check_universe, describe)

print(describe())
print(f"[versions] csdid {_csdid_version}, bootstrap seed {BOOTSTRAP_SEED}")

class SuppressPrint:
    def __enter__(self):
        self._orig = sys.stdout
        sys.stdout = io.StringIO()
        return self

    def __exit__(self, *args):
        sys.stdout = self._orig

# ============================================================
# 1. Load data
# ============================================================
qwi = pd.read_csv(_DIR / 'qwi_all_naics_annual.csv', dtype={'naics': str})
dc = pd.read_csv(_DIR / 'dc_facilities_by_state_year.csv')
mult = pd.read_csv(_DIR / 'pwc_multipliers.csv')

STATE_ABBR_MAP = {
    'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA',
    'Colorado':'CO','Connecticut':'CT','Delaware':'DE','District of Columbia':'DC',
    'Florida':'FL','Georgia':'GA','Hawaii':'HI','Idaho':'ID','Illinois':'IL',
    'Indiana':'IN','Iowa':'IA','Kansas':'KS','Kentucky':'KY','Louisiana':'LA',
    'Maine':'ME','Maryland':'MD','Massachusetts':'MA','Michigan':'MI','Minnesota':'MN',
    'Mississippi':'MS','Missouri':'MO','Montana':'MT','Nebraska':'NE','Nevada':'NV',
    'New Hampshire':'NH','New Jersey':'NJ','New Mexico':'NM','New York':'NY',
    'North Carolina':'NC','North Dakota':'ND','Ohio':'OH','Oklahoma':'OK','Oregon':'OR',
    'Pennsylvania':'PA','Rhode Island':'RI','South Carolina':'SC','South Dakota':'SD',
    'Tennessee':'TN','Texas':'TX','Utah':'UT','Vermont':'VT','Virginia':'VA',
    'Washington':'WA','West Virginia':'WV','Wisconsin':'WI','Wyoming':'WY'
}
ABBR_TO_NAME = {v: k for k, v in STATE_ABBR_MAP.items()}
# The source files spell the District of Columbia "DC".  pwc's
# `state` column uses the same literal, so local_share/local_multiplier are
# keyed on 'DC' too.
STATE_ABBR_MAP['DC'] = 'DC'
ABBR_TO_NAME['DC'] = 'DC'
YEARS = list(range(2016, 2025))

# ============================================================
# 2. Spillover decomposition
# ============================================================
NAT_MULT = mult['emp_total_no_spillover_avg'].sum() / mult['emp_direct_avg'].sum()
mult['spillover_dc'] = mult['emp_spillover_avg'] / NAT_MULT
mult['local_share'] = mult['emp_direct_avg'] / (mult['emp_direct_avg'] + mult['spillover_dc'])

local_share_dict = dict(zip(mult['state'], mult['local_share']))
local_mult_dict = dict(zip(mult['state'], mult['emp_multiplier_local']))

# ============================================================
# 3. Build panels
# ============================================================
dc_states = dc[(dc['state_abbr'] != 'US')
               & ~dc['state_abbr'].isin(OUT_OF_SCOPE_UNITS)].copy()
check_universe(dc_states, 'state_abbr', where='dc capacity file')
dc_panel = []
for _, row in dc_states.iterrows():
    for yr in YEARS:
        dc_panel.append({
            'state_abbr': row['state_abbr'], 'year': yr,
            'dc_capacity_GW': row[f'MW_{yr}'] / 1000
        })
dc_panel = pd.DataFrame(dc_panel)
check_universe(dc_panel, 'state_abbr', where='dc_panel (treatment assignment)')
state_ids = {s: i + 1 for i, s in enumerate(sorted(dc_panel['state_abbr'].unique()))}

qwi['state_abbr'] = qwi['state_name'].map(STATE_ABBR_MAP)
assert qwi['state_abbr'].notna().all(), \
    sorted(qwi.loc[qwi.state_abbr.isna(), 'state_name'].unique())
qwi = qwi[~qwi['state_abbr'].isin(OUT_OF_SCOPE_UNITS)].copy()
check_universe(qwi, 'state_abbr', where='qwi')

# Employment panel
q5182 = qwi[qwi['naics'] == '5182'][['state_abbr', 'year', 'Emp_annual_avg']].rename(
    columns={'Emp_annual_avg': 'emp_5182'})
emp_panel = q5182.merge(dc_panel[['state_abbr', 'year']], on=['state_abbr', 'year'], how='inner')
emp_panel['state_id'] = emp_panel['state_abbr'].map(state_ids)
emp_panel['state_name'] = emp_panel['state_abbr'].map(ABBR_TO_NAME)
emp_panel['local_share'] = emp_panel['state_name'].map(local_share_dict)
emp_panel['local_multiplier'] = emp_panel['state_name'].map(local_mult_dict)
emp_panel['emp_cleaned_direct'] = emp_panel['emp_5182'] * emp_panel['local_share']
emp_panel['emp_cleaned_multiplied'] = emp_panel['emp_cleaned_direct'] * emp_panel['local_multiplier']
assert emp_panel['local_share'].notna().all(), \
    sorted(emp_panel.loc[emp_panel.local_share.isna(), 'state_abbr'].unique())
check_universe(emp_panel, 'state_abbr', where='employment panel')

def report_panel(p, name):
    print(f"[panel {name}] {p['state_abbr'].nunique()} units, {len(p)} unit-year "
          f"obs; per-year n: "
          + " ".join(f"{y}:{p[p.year == y]['state_abbr'].nunique()}" for y in YEARS))

report_panel(emp_panel, 'employment')

# Construction panels
const_panels = {}
for naics in ['23', '2362', '5415']:
    q = qwi[qwi['naics'] == naics][['state_abbr', 'year', 'Emp_annual_avg']].rename(
        columns={'Emp_annual_avg': f'emp_{naics}'})
    cp = q.merge(dc_panel[['state_abbr', 'year']], on=['state_abbr', 'year'], how='inner')
    cp['state_id'] = cp['state_abbr'].map(state_ids)
    check_universe(cp, 'state_abbr', where=f'NAICS {naics} panel')
    report_panel(cp, f'NAICS {naics}')
    const_panels[naics] = cp

# ============================================================
# 4. Treatment assignment
# ============================================================
def get_treatment_years(threshold_gw):
    """Treatment assignment.  Iterates dc_panel, which is now restricted to the
    analysis universe, so the assignment and the outcome panel cannot disagree
    about the sample."""
    treat = {}
    for abbr in dc_panel['state_abbr'].unique():
        st = dc_panel[dc_panel['state_abbr'] == abbr].sort_values('year')
        cap_2016 = st[st['year'] == 2016]['dc_capacity_GW'].values[0]
        crossed = st[(st['dc_capacity_GW'] - cap_2016 >= threshold_gw) & (st['year'] > 2016)]
        treat[abbr] = crossed['year'].iloc[0] if len(crossed) > 0 else 0
    return treat

# ============================================================
# 5. Run all specs
# ============================================================
ALL_SPECS = [
    ('Employment', 'emp_5182', 'Raw NAICS 5182', emp_panel),
    ('Employment', 'emp_cleaned_direct', 'Cleaned Direct', emp_panel),
    ('Employment', 'emp_cleaned_multiplied', 'Cleaned × Multiplier', emp_panel),
    ('Construction', 'emp_23', 'NAICS 23 (Construction)', const_panels['23']),
    ('Construction', 'emp_2362', 'NAICS 2362 (Nonresidential)', const_panels['2362']),
    ('Construction', 'emp_5415', 'NAICS 5415 (Computer Systems)', const_panels['5415']),
]

all_simple = []
all_dynamic = []
all_cohorts = []

for threshold_gw, threshold_label in [(0.5, '500MW'), (1.0, '1GW')]:
    treat_years = get_treatment_years(threshold_gw)
    treated_states = {k: v for k, v in treat_years.items() if v > 0}
    control_states = {k: v for k, v in treat_years.items() if v == 0}
    n_t, n_c = len(treated_states), len(control_states)
    assert n_t + n_c == N_ANALYSIS_UNITS, (n_t, n_c)
    print(f"\n[{threshold_label}] treatment assignment over "
          f"{n_t + n_c} units: {n_t} treated, {n_c} never-treated")

    cohorts = Counter(treated_states.values())
    for yr in sorted(cohorts):
        sts = sorted([k for k, v in treated_states.items() if v == yr])
        all_cohorts.append({
            'threshold': threshold_label, 'cohort_year': yr,
            'n_states': len(sts), 'states': ', '.join(sts)
        })

    for category, y_col, y_label, base_panel in ALL_SPECS:
        df = base_panel[['state_id', 'state_abbr', 'year', y_col]].copy()
        df['treat_year'] = df['state_abbr'].map(treat_years)
        df = df.dropna(subset=[y_col])

        try:
            np.random.seed(BOOTSTRAP_SEED)
            with SuppressPrint():
                model = att_gt.ATTgt(
                    yname=y_col, tname='year', idname='state_id',
                    gname='treat_year', data=df, control_group='nevertreated'
                )
                model.fit()
                # ATTgt.aggte() returns `self`, so the two
                # aggregations share one `atte` dict.  Each result must be
                # copied out BEFORE the next aggregation is requested.  The
                # published code read the simple result after asking for the
                # dynamic one and therefore silently reported the dynamic
                # overall ATT under the "Overall ATT" heading.
                agg_s = model.aggte(typec='simple')
                att_simple = float(agg_s.atte['overall_att'])
                se_simple = float(agg_s.atte['overall_se'].flat[0])

                agg_d = model.aggte(typec='dynamic')
                att_dyn = float(agg_d.atte['overall_att'])
                se_dyn = float(agg_d.atte['overall_se'].flat[0])
                egt = np.array(agg_d.atte['egt'])
                att_egt = np.array(agg_d.atte['att_egt'])
                se_egt = np.array(agg_d.atte['se_egt'])

            assert agg_s is agg_d, \
                "csdid no longer aliases aggte results; revisit this block"

            # `ATT` stays the DYNAMIC overall, which is what the published
            # table actually contains, so the published series remains
            # traceable.  `ATT_simple` is the correctly-labelled simple
            # aggregation, reported alongside it.
            att, se = att_dyn, se_dyn
            p = 2 * (1 - norm.cdf(abs(att / se)))
            p_simple = 2 * (1 - norm.cdf(abs(att_simple / se_simple)))

            all_simple.append({
                'threshold': threshold_label, 'category': category, 'y_var': y_label,
                'ATT': att, 'se': se, 'ci_low': att - 1.96 * se, 'ci_high': att + 1.96 * se,
                'p_approx': p, 'n_treated': n_t, 'n_control': n_c,
                'ATT_simple': att_simple, 'se_simple': se_simple,
                'p_simple': p_simple,
                'n_units': n_t + n_c, 'n_obs': int(len(df)),
            })

            if se_egt.ndim == 2:
                se_egt = se_egt[0]

            for e, a, s in zip(egt, att_egt, se_egt):
                all_dynamic.append({
                    'threshold': threshold_label, 'category': category, 'y_var': y_label,
                    'event_time': int(e), 'ATT': float(a), 'se': float(s),
                    'ci_low': float(a - 1.96 * s), 'ci_high': float(a + 1.96 * s),
                })

            sig = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else ''
            print(f"OK  {threshold_label:5s}  {y_label:30s}  "
                  f"ATT_dyn={att:>10,.1f} (SE {se:>8,.1f}, p {p:.4f}{sig})   "
                  f"ATT_simple={att_simple:>10,.1f} (SE {se_simple:>8,.1f}, "
                  f"p {p_simple:.4f})   N={len(df)}/{n_t + n_c}")
        except Exception as ex:
            print(f"ERR {threshold_label:5s}  {y_label:30s}  {ex}")

print(f"\nTotal: {len(all_simple)} overall, {len(all_dynamic)} event-study, {len(all_cohorts)} cohorts")

# ============================================================
# 6. Write Excel
# ============================================================
pass
pass
pass
pass

wb = Workbook()
wb.remove(wb.active)

# --- Sheet 1: Overall ATT ---
ws = wb.create_sheet("Overall ATT")
pass
ws.freeze_panes = "A3"
pass

ws.merge_cells('A1:N1')
c = ws['A1']
c.value = ("Staggered DiD (Callaway & Sant'Anna 2021) — overall ATT, "
           "dynamic and simple aggregations reported separately")
pass
pass
pass

headers = ['Threshold', 'Category', 'Y Variable', 'ATT\n(dynamic agg)',
           'SE', 'CI95 Lower', 'CI95 Upper', 'p-value', 'Treated', 'Control',
           'ATT\n(simple agg)', 'SE\n(simple)', 'p-value\n(simple)', 'N obs']
pass
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=ci, value=h)
    pass
    pass
    pass
    pass

fmts = [None, None, None, '#,##0.0', '#,##0.0', '#,##0.0', '#,##0.0', '0.0000',
        None, None, '#,##0.0', '#,##0.0', '0.0000', None]
for idx, r in enumerate(all_simple):
    rn = idx + 3
    vals = [r['threshold'], r['category'], r['y_var'], r['ATT'], r['se'],
            r['ci_low'], r['ci_high'], r['p_approx'], r['n_treated'], r['n_control'],
            r['ATT_simple'], r['se_simple'], r['p_simple'], r['n_obs']]
    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws.cell(row=rn, column=ci, value=v)
        pass
        if fmt:
            pass
    if r['p_approx'] < 0.05:
        pass
        for ci in [4, 5, 6, 7, 8]:
            ws.cell(row=rn, column=ci)

sr = len(all_simple) + 4
ws.cell(row=sr, column=1, value="Notes")
for i, (l, v) in enumerate([
    ("Model", "Callaway & Sant'Anna (2021) Doubly Robust DiD"),
    ("Universe", UNIVERSE_LABEL + f" ({N_ANALYSIS_UNITS} units)"),
    ("Treatment", "First year NEW DC capacity (since 2016) > threshold"),
    ("Control", "Never-treated units (never exceeded threshold by 2024)"),
    ("SE", f"Bootstrapped (1000 iterations), seed {BOOTSTRAP_SEED}"),
    ("Period", "2016–2024; Michigan absent 2022–2024"),
    ("ATT (dynamic agg)", "average of the event-study ATT(e) for e >= 0."),
    ("", "This is the column the published table contained, mislabelled"),
    ("", "as the simple aggregation (csdid aggte() returns self)."),
    ("ATT (simple agg)", "group-size-weighted average of all post ATT(g,t)."),
    ("csdid version", f"pinned {REQUIRED_CSDID}; point estimates are not stable"),
    ("", "across csdid versions -- 0.4.2 moves 1 GW raw 5182 by 6.5%"),
]):
    ws.cell(row=sr + 1 + i, column=1, value=l)
    ws.cell(row=sr + 1 + i, column=3, value=v)

for ci, w in enumerate([18, 14, 28, 14, 10, 12, 12, 10, 8, 8, 14, 11, 11, 8], 1):
    pass

# --- Sheet 2: Event Study ---
ws2 = wb.create_sheet("Event Study")
pass
ws2.freeze_panes = "A2"
headers2 = ['Threshold', 'Category', 'Y Variable', 'Event Time', 'ATT', 'SE',
            'CI95 Lower', 'CI95 Upper', 'Sig']
for ci, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    pass
    pass
    pass
    pass

fmts2 = [None, None, None, None, '#,##0.0', '#,##0.0', '#,##0.0', '#,##0.0', None]
for idx, r in enumerate(all_dynamic):
    rn = idx + 2
    se_v = r['se']
    att_v = r['ATT']
    sig = (abs(att_v / se_v) > 1.96) if se_v > 0 else False
    vals = [r['threshold'], r['category'], r['y_var'], r['event_time'],
            att_v, se_v, r['ci_low'], r['ci_high'], 'Yes' if sig else '']
    for ci, (v, fmt) in enumerate(zip(vals, fmts2), 1):
        cell = ws2.cell(row=rn, column=ci, value=v)
        pass
        if fmt:
            pass
    if r['event_time'] < 0 and sig:
        for ci in [5, 6, 9]:
            ws2.cell(row=rn, column=ci)
    elif r['event_time'] >= 0 and sig:
        pass
        for ci in [5, 6, 9]:
            ws2.cell(row=rn, column=ci)

for ci, w in enumerate([10, 14, 28, 10, 12, 10, 12, 12, 6], 1):
    pass

# --- Sheet 3: Treatment Cohorts ---
ws3 = wb.create_sheet("Treatment Cohorts")
pass
ws3.freeze_panes = "A2"
for ci, h in enumerate(['Threshold', 'Cohort Year', '# States', 'States'], 1):
    cell = ws3.cell(row=1, column=ci, value=h)
    pass
    pass
    pass
for idx, r in enumerate(all_cohorts):
    rn = idx + 2
    for ci, v in enumerate([r['threshold'], r['cohort_year'], r['n_states'], r['states']], 1):
        ws3.cell(row=rn, column=ci, value=v)
for ci, w in enumerate([10, 12, 10, 60], 1):
    pass

# --- Sheet 4: Event Study Plot Data ---
ws4 = wb.create_sheet("Event Study Plot Data")
pass
key_specs = [
    ('500MW', 'Cleaned Direct'),
    ('500MW', 'NAICS 2362 (Nonresidential)'),
    ('1GW', 'Cleaned Direct'),
    ('1GW', 'NAICS 2362 (Nonresidential)'),
]
col_idx = 1
for threshold, y_var in key_specs:
    sub = sorted(
        [r for r in all_dynamic if r['threshold'] == threshold and r['y_var'] == y_var],
        key=lambda x: x['event_time']
    )
    if not sub:
        continue
    ws4.cell(row=1, column=col_idx, value=f"{y_var} ({threshold})")
    for ci, h in enumerate(['e', 'ATT', 'CI_lo', 'CI_hi']):
        c = ws4.cell(row=2, column=col_idx + ci, value=h)
        pass
        pass
    for i, r in enumerate(sub):
        rn = i + 3
        for ci, v in enumerate([r['event_time'], r['ATT'], r['ci_low'], r['ci_high']]):
            c = ws4.cell(row=rn, column=col_idx + ci, value=v)
            pass
            if ci > 0:
                pass
    col_idx += 5

wb.save(_DIR / '06_staggered_did_analysis.xlsx')
print(f"\nSaved 06_staggered_did_analysis.xlsx")
print(f"Sheets: {wb.sheetnames}")

_outdir = _DIR / '../results/r6_employment'
pd.DataFrame(all_simple).to_csv(_outdir / 'did_overall_att.csv', index=False)
pd.DataFrame(all_cohorts).to_csv(_outdir / 'did_cohorts.csv', index=False)
pd.DataFrame(all_dynamic).to_csv(_outdir / 'did_event_study.csv', index=False)
print(f"Saved {_outdir / 'did_overall_att.csv'}")
print(f"Saved {_outdir / 'did_cohorts.csv'}")
print(f"Saved {_outdir / 'did_event_study.csv'}")
