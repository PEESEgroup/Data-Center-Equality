"""
Intra-industry substitution analysis: NAICS 5182 vs 517 vs 51
For states with DC capacity > 1GW (2024) + national total.

Compares employment changes (2016→2024) across:
  - NAICS 5182: Data Processing, Hosting & Related
  - NAICS 51:  Information Sector (parent)
  - NAICS 517: Telecommunications

Output: 03_substitution_analysis.xlsx
  Sheet 1: Substitution Summary (side-by-side changes)
  Sheet 2: Time Series (annual employment 2016-2024)
  Sheet 3: 5182 vs 517 Comparison (net analysis & substitution ratio)

IMPLEMENTATION NOTES
-----------------------------------------
The national row of this table turns on two construction choices.

1. UNIVERSE.  National totals are summed over the analysis universe imported
   from `analysis_universe.py` -- the 48 contiguous states plus DC -- not over
   all 51 QWI labels.
2. BALANCE.  A `.sum()` by year over an unbalanced panel counts a unit's exit
   from the data as an employment decline: Alaska is observed only in 2016 and
   Michigan is absent from 2022-2024, so their base-year employment enters the
   2016 total while their 2024 employment does not.  National totals are
   therefore summed over units observed in every year of 2016-2024 via
   `panel_aggregates.py`, which prints its n per year.  The unbalanced series
   is written alongside as a diagnostic so the size of the difference is on
   the record.

Consequence: the 2016-2024 net change in NAICS 5182 + 517 changes SIGN, and the
headline substitution ratio moves from 0.87 to about 1.01.  The 15 per-state
rows are unaffected -- none of them is AK, HI or MI.
"""

import sys
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook

# ============================================================
# 1. Load data
# ============================================================
# Release-set path rule: no absolute paths, and no assumption about the
# working directory.  This script now lives IN employment.
_DIR = Path(__file__).resolve().parent                    # employment
sys.path.insert(0, str(_DIR))
from analysis_universe import (N_ANALYSIS_UNITS, OUT_OF_SCOPE_UNITS,
                               UNIVERSE_LABEL, check_universe, describe)
from panel_aggregates import balanced_sum_by_year

print(describe())

qwi = pd.read_csv(_DIR / 'qwi_all_naics_annual.csv', dtype={'naics': str})
dc = pd.read_csv(_DIR / 'dc_facilities_by_state_year.csv')

STATE_ABBR_MAP = {
    'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA',
    'Colorado':'CO','Connecticut':'CT','Delaware':'DE','District of Columbia':'DC',
    'Florida':'FL','Georgia':'GA','Hawaii':'HI','Idaho':'ID','Illinois':'IL',
    'Indiana':'IN','Iowa':'IA','Kansas':'KS','Kentucky':'KY','Louisiana':'LA',
    'Maine':'ME','Maryland':'MD','Massachusetts':'MA','Michigan':'MI','Minnesota':'MN',
    'Mississippi':'MS','Missouri':'MO','Montana':'MT','Nebraska':'NE','Nevada':'NV',
    'New Hampshire':'NH','New Jersey':'NJ','New Mexico':'NM','New York':'NY',
    'North Carolina':'NC','North Dakota':'ND','Ohio':'OH','Oklahoma':'OK','Oregon':'OR',
    'Pennsylvania':'PA','Rhode Island':'RI','South Carolina':'SC','South Dakota':'SD',
    'Tennessee':'TN','Texas':'TX','Utah':'UT','Vermont':'VT','Virginia':'VA',
    'Washington':'WA','West Virginia':'WV','Wisconsin':'WI','Wyoming':'WY'
}
ABBR_TO_NAME = {v: k for k, v in STATE_ABBR_MAP.items()}
# The source files spell the District of Columbia "DC".
STATE_ABBR_MAP['DC'] = 'DC'
ABBR_TO_NAME['DC'] = 'District of Columbia'
# 2026-07-31: the descriptive comparison now opens at the same year as the causal
# design (SI 6.2), so that the paper does not report a raw co-movement over a window
# the instrument is not identified on.  Nearly half of the 2016-2024 telecommunications
# decline had already happened before 2019, which is itself evidence against
# displacement and was invisible while the base year was 2016.
BASE_YEAR = 2019
YEARS = list(range(BASE_YEAR, 2025))

# ============================================================
# 2. Get 2024 DC capacity and filter states > 1GW
# ============================================================
dc_states = dc[(dc['state_abbr'] != 'US')
               & ~dc['state_abbr'].isin(OUT_OF_SCOPE_UNITS)].copy()
check_universe(dc_states, 'state_abbr', where='dc capacity file')

dc_2024 = dc_states[['state_abbr', 'MW_2024']].copy()
dc_2024['cap_GW'] = dc_2024['MW_2024'] / 1000
dc_2024 = dc_2024[dc_2024['cap_GW'] >= 1.0].sort_values('cap_GW', ascending=False)
big_states = list(dc_2024['state_abbr'])

# National total capacity.  NOT the 'US' row of the capacity file: that row is
# the 51-unit total and includes AK and HI, which are out of scope.  Summed
# over the 49 analysis units instead.
nat_cap = dc_states['MW_2024'].sum() / 1000
nat_cap_file_us_row = dc[dc['state_abbr'] == 'US']['MW_2024'].values[0] / 1000
print(f"[capacity] 2024 total over {len(dc_states)} analysis units = "
      f"{nat_cap:.6f} GW; the capacity file's 51-unit 'US' row = "
      f"{nat_cap_file_us_row:.6f} GW (difference = AK + HI = "
      f"{nat_cap_file_us_row - nat_cap:.6f} GW)")

# ============================================================
# 3. Build employment pivots for 3 NAICS codes
# ============================================================
qwi['state_abbr'] = qwi['state_name'].map(STATE_ABBR_MAP)
assert qwi['state_abbr'].notna().all(), \
    sorted(qwi.loc[qwi.state_abbr.isna(), 'state_name'].unique())
qwi = qwi[~qwi['state_abbr'].isin(OUT_OF_SCOPE_UNITS)].copy()
check_universe(qwi, 'state_abbr', where='qwi')

pivots = {}
for naics in ['5182', '51', '517']:
    q = qwi[qwi['naics'] == naics].copy()
    piv = q.pivot_table(index='state_abbr', columns='year', values='Emp_annual_avg')
    pivots[naics] = piv

# ---------------------------------------------------------------------------
# National totals.  A plain `groupby('year').sum()` over an
# unbalanced 51-unit panel.  Both faults are fixed here.
# ---------------------------------------------------------------------------
nat_totals = {}
nat_totals_unbalanced = {}
nat_meta = {}
print("\n[national totals] balanced over units observed in every year of the window")
for naics in ['5182', '51', '517']:
    q = qwi[qwi['naics'] == naics]
    tot, meta = balanced_sum_by_year(
        q, 'state_abbr', 'year', 'Emp_annual_avg', years=YEARS,
        label=f'national NAICS {naics} employment')
    nat_totals[naics] = tot
    nat_totals_unbalanced[naics] = meta['unbalanced_total']
    nat_meta[naics] = meta

_bal_units = nat_meta['5182']['balanced_units']
assert all(nat_meta[n]['balanced_units'] == _bal_units for n in nat_totals), \
    "the three NAICS series must be balanced on the same set of units"
print(f"[national totals] all three series on the SAME {len(_bal_units)} "
      f"balanced units; excluded from the national row: "
      f"{nat_meta['5182']['dropped_units'] or 'none'}")

# Capacity of the balanced set, so the national row can be read either way.
nat_cap_balanced = (dc_states[dc_states['state_abbr'].isin(_bal_units)]['MW_2024']
                    .sum() / 1000)
print(f"[capacity] 2024 total over the {len(_bal_units)} BALANCED units = "
      f"{nat_cap_balanced:.6f} GW")

# ============================================================
# 4. Build summary table
# ============================================================
rows = []
for abbr in big_states:
    st_name = ABBR_TO_NAME.get(abbr, abbr)
    cap = dc_2024[dc_2024['state_abbr'] == abbr]['cap_GW'].values[0]
    row = {'state_abbr': abbr, 'state_name': st_name, 'cap_GW': cap}

    for naics, prefix in [('5182', 'dc'), ('51', 'info'), ('517', 'tel')]:
        if abbr in pivots[naics].index:
            e16 = pivots[naics].loc[abbr, BASE_YEAR] if BASE_YEAR in pivots[naics].columns else np.nan
            e24 = pivots[naics].loc[abbr, 2024] if 2024 in pivots[naics].columns else np.nan
        else:
            e16 = e24 = np.nan
        delta = e24 - e16 if not (np.isnan(e16) or np.isnan(e24)) else np.nan
        pct = delta / e16 if (e16 and not np.isnan(delta)) else np.nan
        row[f'{prefix}_{BASE_YEAR}'] = e16
        row[f'{prefix}_2024'] = e24
        row[f'{prefix}_delta'] = delta
        row[f'{prefix}_pct'] = pct
    rows.append(row)

# National row -- balanced, 49-unit universe
nat_row = {'state_abbr': 'US', 'state_name': 'National', 'cap_GW': nat_cap}
for naics, prefix in [('5182', 'dc'), ('51', 'info'), ('517', 'tel')]:
    e16 = nat_totals[naics].get(BASE_YEAR, np.nan)
    e24 = nat_totals[naics].get(2024, np.nan)
    delta = e24 - e16
    nat_row[f'{prefix}_{BASE_YEAR}'] = e16
    nat_row[f'{prefix}_2024'] = e24
    nat_row[f'{prefix}_delta'] = delta
    nat_row[f'{prefix}_pct'] = delta / e16 if e16 else np.nan
rows.append(nat_row)

summary = pd.DataFrame(rows)

# ---------------------------------------------------------------------------
# National-row diagnostics: what each basis gives.  Written out so the SI stage
# never has to guess which basis a published number came from.
# ---------------------------------------------------------------------------
def _nat_block(series_by_naics, label, n_units_note):
    d = {'basis': label, 'units_note': n_units_note}
    for naics, prefix in [('5182', 'dc'), ('51', 'info'), ('517', 'tel')]:
        s = series_by_naics[naics]
        e16, e24 = float(s.get(BASE_YEAR)), float(s.get(2024))
        d[f'{prefix}_{BASE_YEAR}'] = e16
        d[f'{prefix}_2024'] = e24
        d[f'{prefix}_delta'] = e24 - e16
        d[f'{prefix}_pct'] = (e24 - e16) / e16
    d['net_5182_plus_517'] = d['dc_delta'] + d['tel_delta']
    d['sub_ratio'] = abs(d['dc_delta'] / d['tel_delta'])
    return d

_diag = [
    _nat_block(nat_totals, f'PRIMARY: 49-unit universe, balanced {BASE_YEAR}-2024',
               f"{len(_bal_units)} units summed in every year "
               f"(excluded: {', '.join(nat_meta['5182']['dropped_units']) or 'none'})"),
    _nat_block(nat_totals_unbalanced, 'diagnostic: 49-unit universe, UNBALANCED',
               'per-year n = ' + " ".join(
                   f"{y}:{nat_meta['5182']['n_raw'].get(y, 0)}" for y in YEARS)),
]
_diag_df = pd.DataFrame(_diag)
_diag_df['cap_GW_49unit'] = nat_cap
_diag_df['cap_GW_balanced'] = nat_cap_balanced
_diag_df['cap_GW_file_US_row_51unit'] = nat_cap_file_us_row
print("\n[national row] basis comparison")
for r in _diag:
    print(f"  {r['basis']}")
    print(f"    {r['units_note']}")
    print(f"    d5182 {r['dc_delta']:>+12,.1f} ({r['dc_pct']:+.1%})   "
          f"d517 {r['tel_delta']:>+12,.1f} ({r['tel_pct']:+.1%})   "
          f"d51 {r['info_delta']:>+12,.1f} ({r['info_pct']:+.1%})")
    print(f"    net 5182+517 {r['net_5182_plus_517']:>+12,.1f}   "
          f"substitution ratio {r['sub_ratio']:.4f}")

# Interpretation columns
def interpret_pattern(r):
    dc_d = r['dc_delta']
    info_d = r['info_delta']
    if pd.isna(dc_d) or pd.isna(info_d):
        return ''
    if dc_d > 0 and info_d < 0:
        return 'Substitution'
    elif dc_d > 0 and info_d > 0:
        if dc_d > info_d:
            return 'Substitution'
        else:
            return 'Net new jobs'
    elif dc_d < 0:
        return 'DC declining'
    return 'Net new jobs'

def interpret_tel(r):
    tel_pct = r['tel_pct']
    if pd.isna(tel_pct):
        return ''
    if tel_pct < -0.1:
        return 'Large decline'
    elif tel_pct < 0:
        return 'Decline'
    return 'Stable/Growth'

summary['pattern'] = summary.apply(interpret_pattern, axis=1)
summary['tel_interp'] = summary.apply(interpret_tel, axis=1)

# ============================================================
# 5. Write Excel
# ============================================================
pass
pass

wb = Workbook()
wb.remove(wb.active)

# --- Sheet 1: Substitution Summary ---
ws = wb.create_sheet("Substitution Summary")
pass
ws.freeze_panes = "A3"
pass

groups = [
    ('A1:C1', 'State', '4472C4'),
    ('D1:G1', 'NAICS 5182 — Data Processing & Hosting', '2F5496'),
    ('H1:K1', 'NAICS 51 — Information Sector', '548235'),
    ('L1:O1', 'NAICS 517 — Telecommunications', 'BF8F00'),
    ('P1:Q1', 'Interpretation', '7030A0'),
]
for rng, title, color in groups:
    ws.merge_cells(rng)
    c = ws[rng.split(':')[0]]
    c.value = title
    pass
    pass

headers = ['State', 'State Name', 'DC Cap\n2024 (GW)',
           f'Emp\n{BASE_YEAR}', 'Emp\n2024', 'Δ Change', '% Change',
           f'Emp\n{BASE_YEAR}', 'Emp\n2024', 'Δ Change', '% Change',
           f'Emp\n{BASE_YEAR}', 'Emp\n2024', 'Δ Change', '% Change',
           '5182 vs 51\nPattern', '517\nDecline']
fills = ['D6E4F0'] * 3 + ['D6E4F0'] * 4 + ['E2EFDA'] * 4 + ['FFF2CC'] * 4 + ['E8D5F5'] * 2
for ci, (h, fc) in enumerate(zip(headers, fills), 1):
    cell = ws.cell(row=2, column=ci, value=h)
    pass
    pass
    pass
    pass
ws.row_dimensions[2].height = 35

fmts = [None, None, '0.000',
        '#,##0', '#,##0', '+#,##0;-#,##0', '+0.0%;-0.0%',
        '#,##0', '#,##0', '+#,##0;-#,##0', '+0.0%;-0.0%',
        '#,##0', '#,##0', '+#,##0;-#,##0', '+0.0%;-0.0%',
        None, None]
keys = ['state_abbr', 'state_name', 'cap_GW',
        f'dc_{BASE_YEAR}', 'dc_2024', 'dc_delta', 'dc_pct',
        f'info_{BASE_YEAR}', 'info_2024', 'info_delta', 'info_pct',
        f'tel_{BASE_YEAR}', 'tel_2024', 'tel_delta', 'tel_pct',
        'pattern', 'tel_interp']

for idx, (_, r) in enumerate(summary.iterrows()):
    rn = idx + 3
    for ci, (k, fmt) in enumerate(zip(keys, fmts), 1):
        v = r[k]
        if pd.isna(v):
            v = None
        cell = ws.cell(row=rn, column=ci, value=v)
        pass
        if fmt:
            pass

for ci, w in enumerate([5, 16, 10, 10, 10, 10, 9, 10, 10, 10, 9, 10, 10, 10, 9, 14, 12], 1):
    pass

# --- Sheet 2: Time Series ---
ws2 = wb.create_sheet("Time Series")
pass
ws2.freeze_panes = "A3"

col_offset = 0
for naics, title, color in [('5182', 'NAICS 5182 — Data Processing & Hosting', '2F5496'),
                              ('51', 'NAICS 51 — Information Sector', '548235'),
                              ('517', 'NAICS 517 — Telecommunications', 'BF8F00')]:
    start_col = col_offset + 1
    end_col = col_offset + 2 + len(YEARS)

    ws2.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    c = ws2.cell(row=1, column=start_col, value=title)
    pass
    pass

    for ci, h in enumerate(['State', 'State Name'] + [str(y) for y in YEARS], start_col):
        cell = ws2.cell(row=2, column=ci, value=h)
        pass
        pass
        pass

    for idx, abbr in enumerate(big_states + ['US']):
        rn = idx + 3
        ws2.cell(row=rn, column=start_col, value=abbr)
        _nm = (f'National (balanced, {len(_bal_units)} units)' if abbr == 'US'
               else ABBR_TO_NAME.get(abbr, abbr))
        ws2.cell(row=rn, column=start_col + 1, value=_nm)

        if abbr == 'US':
            for yi, yr in enumerate(YEARS):
                v = nat_totals[naics].get(yr, None)
                cell = ws2.cell(row=rn, column=start_col + 2 + yi, value=v)
                pass
        elif abbr in pivots[naics].index:
            for yi, yr in enumerate(YEARS):
                v = pivots[naics].loc[abbr, yr] if yr in pivots[naics].columns else None
                cell = ws2.cell(row=rn, column=start_col + 2 + yi, value=v)
                pass

    col_offset = end_col + 1

# --- Sheet 3: 5182 vs 517 Comparison ---
ws3 = wb.create_sheet("5182 vs 517 Comparison")
pass
ws3.freeze_panes = "A3"

groups3 = [
    ('A1:C1', 'State', '4472C4'),
    ('D1:E1', 'NAICS 5182 (DC)', '2F5496'),
    ('F1:G1', 'NAICS 517 (Telecom)', 'BF8F00'),
    ('H1:I1', 'NAICS 51 (Info)', '548235'),
    ('J1:K1', 'Net Analysis', 'C00000'),
]
for rng, title, color in groups3:
    ws3.merge_cells(rng)
    c = ws3[rng.split(':')[0]]
    c.value = title
    pass
    pass

headers3 = ['State', 'State Name', 'DC Cap\n(GW)',
            'Δ 5182', '% Chg', 'Δ 517', '% Chg', 'Δ 51', '% Chg',
            '5182+517\nNet', 'Substitution\nRatio']
for ci, h in enumerate(headers3, 1):
    cell = ws3.cell(row=2, column=ci, value=h)
    pass
    pass
    pass
ws3.row_dimensions[2].height = 35

fmts3 = [None, None, '0.000',
         '+#,##0;-#,##0', '+0.0%;-0.0%',
         '+#,##0;-#,##0', '+0.0%;-0.0%',
         '+#,##0;-#,##0', '+0.0%;-0.0%',
         '+#,##0;-#,##0', '0.00']

for idx, (_, r) in enumerate(summary.iterrows()):
    rn = idx + 3
    net = (r['dc_delta'] or 0) + (r['tel_delta'] or 0)
    sub_ratio = abs(r['dc_delta'] / r['tel_delta']) if (r['tel_delta'] and r['tel_delta'] != 0) else np.nan

    vals = [r['state_abbr'], r['state_name'], r['cap_GW'],
            r['dc_delta'], r['dc_pct'], r['tel_delta'], r['tel_pct'],
            r['info_delta'], r['info_pct'], net, sub_ratio]
    for ci, (v, fmt) in enumerate(zip(vals, fmts3), 1):
        if pd.isna(v):
            v = None
        cell = ws3.cell(row=rn, column=ci, value=v)
        pass
        if fmt:
            pass

for ci, w in enumerate([5, 16, 8, 10, 8, 10, 8, 10, 8, 10, 10], 1):
    pass

wb.save(_DIR / '03_substitution_analysis.xlsx')
print(f"\nSaved 03_substitution_analysis.xlsx")
print(f"Units > 1GW: {len(big_states)} + national = {len(big_states)+1} rows")

_outdir = _DIR / '../results/r6_employment'
_diag_df.to_csv(_outdir / 'substitution_national.csv', index=False)
summary.to_csv(_outdir / 'substitution_table.csv', index=False)
pd.DataFrame({'year': YEARS,
              **{f'n_units_naics_{n}': [nat_meta[n]['n_raw'].get(y, 0) for y in YEARS]
                 for n in ['5182', '51', '517']},
              **{f'balanced_n_naics_{n}': [nat_meta[n]['n_balanced'].get(y, 0) for y in YEARS]
                 for n in ['5182', '51', '517']},
              **{f'total_naics_{n}': [float(nat_totals[n].get(y)) for y in YEARS]
                 for n in ['5182', '51', '517']},
              **{f'total_unbal_naics_{n}': [float(nat_totals_unbalanced[n].get(y)) for y in YEARS]
                 for n in ['5182', '51', '517']},
              }).to_csv(_outdir / 'substitution_year_counts.csv', index=False)
print(f"Saved {_outdir / 'substitution_national.csv'}")
print(f"Saved {_outdir / 'substitution_table.csv'}")
print(f"Saved {_outdir / 'substitution_year_counts.csv'}")
