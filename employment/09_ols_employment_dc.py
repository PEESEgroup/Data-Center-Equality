"""
OLS: Employment = α + β × DC_capacity(GW) + ε
State-level, 2016-2024.

Four employment specifications:
  1. Raw NAICS 5182 (observed direct)
  2. Raw NAICS 5182 × local_multiplier + spillover (double-counting issue)
  3. Cleaned direct = NAICS 5182 × local_share (removing spillover-driven 5182 jobs)
  4. Cleaned direct × local_multiplier (total local impact, no double-counting)

Spillover decomposition method:
  - National multiplier = Σ(total_no_spillover) / Σ(direct) ≈ 5.43
  - DC industry share of total = 1/national_multiplier ≈ 18%
  - spillover_dc_i = PwC_spillover_i / national_multiplier
  - local_share_i = PwC_direct_i / (PwC_direct_i + spillover_dc_i)

Outputs: 02_ols_employment_dc_capacity.xlsx
  Sheet 1: OLS by Capacity (raw)
  Sheet 2: Panel Data
  Sheet 3: Cleaned OLS by Capacity
  Sheet 4: Panel FE (State + Year fixed effects, clustered SE)

IMPLEMENTATION NOTES
-------------------------------------------------
1. UNIVERSE.  The analysis universe is imported from `employment/analysis_universe.py`
   and is never re-declared here.  It is the 48 contiguous states plus DC, 49 units.
   The published run was the exact complement of that at both ends: AK and HI were
   retained (AK contributing a single 2016 observation) and DC was silently dropped,
   because STATE_ABBR_MAP carried only 'District of Columbia'->'DC' while
   qwi_all_naics_annual.csv and pwc_multipliers.csv both spell the unit as the
   literal two-letter string "DC", so "DC".map(...) returned NaN.
2. BALANCED AGGREGATES.  `direct_by_year`, the national NAICS 5182 total that enters
   the spec-2 spillover term, was a `.sum()` over an unbalanced panel: Michigan
   leaves QWI after 2021, so the 2022-2024 national totals silently lost ~8,570
   jobs and every state's spec-2 outcome inherited a step change that is an
   artefact of the data, not of employment.  It is now summed over units observed
   in every year of 2016-2024, via `employment/panel_aggregates.py`, which prints
   its n per year.  The unbalanced variant is still computed and reported as a
   diagnostic so the size of the artefact is on the record.
3. MIN_OBS_YEARS.  A unit with fewer than 3 observed years cannot enter the state
   OLS or be counted in "N units".  On the 49-unit universe nothing is excluded by
   this guard (Michigan has 6 years); it exists so a one-observation unit like the
   published run's Alaska can never be counted again.
4. SE CONVENTION.  linearmodels reports the uncorrected clustered SE.  SI section 6
   prints the Cameron-Gelbach-Miller finite-sample-corrected SE with t(G-1)
   inference.  Both are now written to the Panel FE sheet, side by side, so the
   two conventions cannot silently diverge across script boundaries again.
"""

import sys
import pandas as pd
import numpy as np
from pathlib import Path
from scipy import stats
from linearmodels.panel import PanelOLS
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ============================================================
# 1. Load data
# ============================================================
# Release-set path rule: no absolute paths, and no assumption about the
# working directory.  This script now lives IN employment.
_DIR = Path(__file__).resolve().parent                    # employment
sys.path.insert(0, str(_DIR))
from analysis_universe import (ANALYSIS_UNITS, N_ANALYSIS_UNITS,
                               OUT_OF_SCOPE_UNITS, UNIVERSE_LABEL,
                               check_universe, describe)
from panel_aggregates import balanced_sum_by_year

MIN_OBS_YEARS = 3

print(describe())

qwi = pd.read_csv(_DIR / 'qwi_all_naics_annual.csv', dtype={'naics': str})
dc = pd.read_csv(_DIR / 'dc_facilities_by_state_year.csv')
mult = pd.read_csv(_DIR / 'pwc_multipliers.csv')

STATE_ABBR_MAP = {
    'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA',
    'Colorado':'CO','Connecticut':'CT','Delaware':'DE','District of Columbia':'DC',
    'Florida':'FL','Georgia':'GA','Hawaii':'HI','Idaho':'ID','Illinois':'IL',
    'Indiana':'IN','Iowa':'IA','Kansas':'KS','Kentucky':'KY','Louisiana':'LA',
    'Maine':'ME','Maryland':'MD','Massachusetts':'MA','Michigan':'MI','Minnesota':'MN',
    'Mississippi':'MS','Missouri':'MO','Montana':'MT','Nebraska':'NE','Nevada':'NV',
    'New Hampshire':'NH','New Jersey':'NJ','New Mexico':'NM','New York':'NY',
    'North Carolina':'NC','North Dakota':'ND','Ohio':'OH','Oklahoma':'OK','Oregon':'OR',
    'Pennsylvania':'PA','Rhode Island':'RI','South Carolina':'SC','South Dakota':'SD',
    'Tennessee':'TN','Texas':'TX','Utah':'UT','Vermont':'VT','Virginia':'VA',
    'Washington':'WA','West Virginia':'WV','Wisconsin':'WI','Wyoming':'WY'
}
# The source files spell the District of Columbia "DC", not
# "District of Columbia".  Without this line the map returns NaN and DC is
# dropped by the inner join, silently.
STATE_ABBR_MAP['DC'] = 'DC'

YEARS = list(range(2016, 2025))
# The Bartik estimation window opens at the 2019 exposure-share baseline; the spillover
# specification produced here is cross-checked against that module by
# 14_panel_fe_table.py, so it is also estimated on the matched window.
IV_YEAR_MIN = 2019
YEARS_MATCHED = list(range(IV_YEAR_MIN, 2025))

# ============================================================
# 2. National multiplier and spillover decomposition
# ============================================================
NAT_MULT = mult['emp_total_no_spillover_avg'].sum() / mult['emp_direct_avg'].sum()  # ≈ 5.43

mult['spillover_dc'] = mult['emp_spillover_avg'] / NAT_MULT
mult['local_share'] = mult['emp_direct_avg'] / (mult['emp_direct_avg'] + mult['spillover_dc'])

local_share_dict = dict(zip(mult['state'], mult['local_share']))
local_mult_dict = dict(zip(mult['state'], mult['emp_multiplier_local']))
spillover_rate_dict = dict(zip(mult['state'], mult['emp_spillover_rate']))

# ============================================================
# 3. Build panel: state × year
# ============================================================
qwi_5182 = qwi[qwi['naics'] == '5182'].copy()
qwi_5182['state_abbr'] = qwi_5182['state_name'].map(STATE_ABBR_MAP)
assert qwi_5182['state_abbr'].notna().all(), \
    sorted(qwi_5182.loc[qwi_5182.state_abbr.isna(), 'state_name'].unique())

# --- scope: drop the out-of-scope units BEFORE anything is computed ---------
qwi_5182 = qwi_5182[~qwi_5182['state_abbr'].isin(OUT_OF_SCOPE_UNITS)].copy()
check_universe(qwi_5182, 'state_abbr', where='qwi NAICS 5182')

dc_states = dc[(dc['state_abbr'] != 'US')
               & ~dc['state_abbr'].isin(OUT_OF_SCOPE_UNITS)].copy()
check_universe(dc_states, 'state_abbr', where='dc capacity file')
dc_panel = []
for _, row in dc_states.iterrows():
    for yr in YEARS:
        dc_panel.append({
            'state_abbr': row['state_abbr'],
            'year': yr,
            'dc_count': row[f'count_{yr}'],
            'dc_capacity_GW': row[f'MW_{yr}'] / 1000
        })
dc_panel = pd.DataFrame(dc_panel)

# --- the national NAICS 5182 total by year, balanced ------------------------
# This series enters spec 2 as `other_direct`, multiplied by a state-specific
# spillover rate, so a step change in it is NOT absorbed by the year fixed
# effect.  Michigan leaves QWI after 2021.  Summed raw, the series would drop
# ~8,570 jobs in 2022 purely because Michigan stopped being reported.  Summed
# over the units observed in every year of 2016-2024 it does not.
print("\n[direct_by_year] national NAICS 5182 total entering the spec-2 "
      "spillover term")
_dbY, _dbY_meta = balanced_sum_by_year(
    qwi_5182, 'state_abbr', 'year', 'Emp_annual_avg', years=YEARS,
    label='national NAICS 5182 employment total (spec-2 spillover base)')
direct_by_year = _dbY.to_dict()
direct_by_year_unbalanced = _dbY_meta['unbalanced_total'].to_dict()
for _y in YEARS:
    print(f"    {_y}  balanced {direct_by_year[_y]:>12,.1f}   "
          f"unbalanced {direct_by_year_unbalanced[_y]:>12,.1f}   "
          f"diff {direct_by_year[_y] - direct_by_year_unbalanced[_y]:>+10,.1f}")

panel = qwi_5182[['state_abbr', 'state_name', 'year', 'Emp_annual_avg']].merge(
    dc_panel, on=['state_abbr', 'year'], how='inner'
)
panel = panel.rename(columns={'Emp_annual_avg': 'emp_5182'})
check_universe(panel, 'state_abbr', where='estimation panel')

# MIN_OBS_YEARS guard: a unit with fewer than MIN_OBS_YEARS observed years is
# perfectly absorbed by its own fixed effect and must not be counted as a unit.
_yrs_per_unit = panel.groupby('state_abbr')['year'].nunique()
_thin = sorted(_yrs_per_unit[_yrs_per_unit < MIN_OBS_YEARS].index)
if _thin:
    print(f"[MIN_OBS_YEARS] dropping {_thin} (< {MIN_OBS_YEARS} observed years)")
    panel = panel[~panel['state_abbr'].isin(_thin)].copy()
else:
    print(f"[MIN_OBS_YEARS] no unit has fewer than {MIN_OBS_YEARS} observed "
          f"years; nothing dropped")
print(f"[panel] {panel['state_abbr'].nunique()} units, {len(panel)} unit-year "
      f"observations; per-year n: "
      + " ".join(f"{y}:{panel[panel.year == y]['state_abbr'].nunique()}"
                 for y in YEARS))

# Raw multiplied (spec 2, has double-counting)
def calc_raw_multiplied(row):
    st = row['state_name']
    if st not in local_mult_dict:
        return np.nan
    direct_i = row['emp_5182']
    other_direct = direct_by_year.get(row['year'], 0) - direct_i
    return direct_i * local_mult_dict[st] + other_direct * spillover_rate_dict[st]

panel['emp_raw_multiplied'] = panel.apply(calc_raw_multiplied, axis=1)

# State total: raw*(lambda*m + 1 - lambda).  Counts the
# cross-state spillover once, at its own level, and contains no national
# aggregate, so the balanced/unbalanced question does not arise for it.
panel['emp_state_total'] = panel['emp_5182'] * (
    panel['state_name'].map(local_share_dict) * panel['state_name'].map(local_mult_dict)
    + 1 - panel['state_name'].map(local_share_dict))

# Cleaned direct (spec 3) and cleaned multiplied (spec 4)
panel['local_share'] = panel['state_name'].map(local_share_dict)
panel['local_multiplier'] = panel['state_name'].map(local_mult_dict)
panel['emp_cleaned_direct'] = panel['emp_5182'] * panel['local_share']
panel['emp_cleaned_multiplied'] = panel['emp_cleaned_direct'] * panel['local_multiplier']

# ============================================================
# 4. OLS regression
# ============================================================
def run_ols(panel_df, x_col, y_direct_col, y_mult_col, extra_cols=None):
    results = []
    for st_abbr in sorted(panel_df['state_abbr'].unique()):
        df_st = panel_df[panel_df['state_abbr'] == st_abbr].dropna(subset=[y_direct_col, x_col])
        st_name = df_st['state_name'].iloc[0]
        n = len(df_st)
        x = df_st[x_col].values.astype(float)

        base = {'state_abbr': st_abbr, 'state_name': st_name, 'n_obs': n,
                'x_min': x.min() if len(x) > 0 else np.nan,
                'x_max': x.max() if len(x) > 0 else np.nan}

        if extra_cols:
            for ec in extra_cols:
                base[ec] = df_st[ec].iloc[0] if ec in df_st.columns else np.nan

        if n < 3 or np.std(x) < 1e-10:
            for pf in ['direct_', 'mult_']:
                for f in ['slope','se','ci95_lower','ci95_upper','intercept','r_squared','p_value']:
                    base[pf+f] = np.nan
            results.append(base)
            continue

        for pf, yc in [('direct_', y_direct_col), ('mult_', y_mult_col)]:
            y = df_st[yc].values
            sl, ic, r, p, se = stats.linregress(x, y)
            ci = 1.96 * se
            base.update({pf+'slope': sl, pf+'se': se, pf+'ci95_lower': sl-ci,
                         pf+'ci95_upper': sl+ci, pf+'intercept': ic,
                         pf+'r_squared': r**2, pf+'p_value': p})
        results.append(base)
    return pd.DataFrame(results)

# Raw regressions (specs 1 & 2)
res_raw_cap = run_ols(panel, 'dc_capacity_GW', 'emp_5182', 'emp_raw_multiplied')

# Cleaned regressions (specs 3 & 4)
res_cln_cap = run_ols(panel, 'dc_capacity_GW', 'emp_cleaned_direct', 'emp_cleaned_multiplied',
                       extra_cols=['local_share', 'local_multiplier'])

# ============================================================
# 5. Write Excel
# ============================================================
pass
pass
pass
pass

def write_raw_sheet(wb, name, tab_color, res_df, x_label, x_unit, x_fmt,
                    hdr1, hdr2, sub1, sub2, slope_fmt):
    ws = wb.create_sheet(name)
    pass
    ws.freeze_panes = "A3"
    pass

    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = "State Info"
    pass
    ws.merge_cells('F1:L1')
    c = ws['F1']; c.value = f"NAICS 5182 OLS (by {x_label})"
    pass
    ws.merge_cells('M1:S1')
    c = ws['M1']; c.value = f"Raw Multiplier+Spillover OLS (by {x_label})"
    pass

    cols = [('A','State'),('B','State Name'),('C','N'),
            ('D',f'{x_label}\nMin'),('E',f'{x_label}\nMax'),
            ('F',f'β\n(jobs/{x_unit})'),('G','SE'),('H','CI95\nLower'),('I','CI95\nUpper'),
            ('J','Intercept'),('K','R²'),('L','p-value'),
            ('M',f'β\n(jobs/{x_unit})'),('N','SE'),('O','CI95\nLower'),('P','CI95\nUpper'),
            ('Q','Intercept'),('R','R²'),('S','p-value')]

    pass
    pass
    pass
    for cl, t in cols:
        cell = ws[f'{cl}2']; cell.value = t
        pass
        pass
        pass
    ws.row_dimensions[2].height = 35

    for idx, (_, r) in enumerate(res_df.sort_values('direct_slope', ascending=False, na_position='last').iterrows()):
        rn = idx + 3
        ws.cell(row=rn,column=1,value=r['state_abbr'])
        ws.cell(row=rn,column=2,value=r['state_name'])
        c = ws.cell(row=rn,column=3,value=int(r['n_obs']))
        for ci, v in [(4,r['x_min']),(5,r['x_max'])]:
            c = ws.cell(row=rn,column=ci,value=v if not pd.isna(v) else None)
        for pf, cs in [('direct_',6),('mult_',13)]:
            for j, (f, fmt) in enumerate(zip(['slope','se','ci95_lower','ci95_upper','intercept','r_squared','p_value'],
                                              [slope_fmt,slope_fmt,slope_fmt,slope_fmt,'#,##0','0.000','0.0000'])):
                v = r[pf+f]
                c = ws.cell(row=rn,column=cs+j,value=v if not pd.isna(v) else None)
        if not pd.isna(r['direct_p_value']) and r['direct_p_value'] < 0.05:
            pass
            ws.cell(row=rn,column=ci)
        if not pd.isna(r['mult_p_value']) and r['mult_p_value'] < 0.05:
            pass
            ws.cell(row=rn,column=ci)

    sr = rn + 2
    ws.cell(row=sr,column=1,value="Summary")
    valid = res_df.dropna(subset=['direct_slope'])
    for i, (l, v) in enumerate([
        ("Valid regressions", f"{len(valid)} / {len(res_df)}"),
        (f"Median β NAICS 5182 (jobs/{x_unit})", f"{valid['direct_slope'].median():,.1f}"),
        (f"Median β raw multiplied (jobs/{x_unit})", f"{valid['mult_slope'].median():,.1f}"),
        ("p<0.05 (NAICS 5182)", f"{(valid['direct_p_value']<0.05).sum()} / {len(valid)}"),
        ("p<0.05 (raw multiplied)", f"{(valid['mult_p_value']<0.05).sum()} / {len(valid)}"),
        ("",""),("Model", f"Employment = α + β × {x_label} + ε"),
        ("Note", "Raw multiplied has double-counting (see Cleaned sheets)"),
    ]):
        ws.cell(row=sr+1+i,column=1,value=l)
        ws.cell(row=sr+1+i,column=3,value=v)

    for col, w in {'A':7,'B':22,'C':5,'D':9,'E':9,'F':13,'G':10,'H':13,'I':13,'J':12,'K':8,'L':9,
                   'M':13,'N':10,'O':13,'P':13,'Q':12,'R':8,'S':9}.items():
        pass

def write_cleaned_sheet(wb, name, tab_color, res_df, x_label, x_unit, x_fmt,
                        hdr1, hdr2, slope_fmt):
    ws = wb.create_sheet(name)
    pass
    ws.freeze_panes = "A3"
    pass

    ws.merge_cells('A1:G1')
    c = ws['A1']; c.value = "State Info"
    pass
    ws.merge_cells('H1:N1')
    c = ws['H1']; c.value = f"Cleaned Direct OLS (by {x_label})"
    pass
    ws.merge_cells('O1:U1')
    c = ws['O1']; c.value = f"Cleaned × Local Multiplier (by {x_label})"
    pass

    cols = [('A','State'),('B','State Name'),('C','N'),
            ('D','Local\nShare'),('E','Local\nMult'),
            ('F',f'{x_label}\nMin'),('G',f'{x_label}\nMax'),
            ('H',f'β\n(jobs/{x_unit})'),('I','SE'),('J','CI95\nLower'),('K','CI95\nUpper'),
            ('L','Intercept'),('M','R²'),('N','p-value'),
            ('O',f'β\n(jobs/{x_unit})'),('P','SE'),('Q','CI95\nLower'),('R','CI95\nUpper'),
            ('S','Intercept'),('T','R²'),('U','p-value')]

    pass
    pass
    pass
    for cl, t in cols:
        cell = ws[f'{cl}2']; cell.value = t
        pass
        pass
        pass
    ws.row_dimensions[2].height = 35

    for idx, (_, r) in enumerate(res_df.sort_values('direct_slope', ascending=False, na_position='last').iterrows()):
        rn = idx + 3
        ws.cell(row=rn,column=1,value=r['state_abbr'])
        ws.cell(row=rn,column=2,value=r['state_name'])
        c = ws.cell(row=rn,column=3,value=int(r['n_obs']))
        c = ws.cell(row=rn,column=4,value=r.get('local_share'))
        c = ws.cell(row=rn,column=5,value=r.get('local_multiplier'))
        for ci, v in [(6,r['x_min']),(7,r['x_max'])]:
            c = ws.cell(row=rn,column=ci,value=v if not pd.isna(v) else None)
        for pf, cs in [('direct_',8),('mult_',15)]:
            for j, (f, fmt) in enumerate(zip(['slope','se','ci95_lower','ci95_upper','intercept','r_squared','p_value'],
                                              [slope_fmt,slope_fmt,slope_fmt,slope_fmt,'#,##0','0.000','0.0000'])):
                v = r[pf+f]
                c = ws.cell(row=rn,column=cs+j,value=v if not pd.isna(v) else None)
        if not pd.isna(r['direct_p_value']) and r['direct_p_value'] < 0.05:
            pass
            ws.cell(row=rn,column=ci)
        if not pd.isna(r['mult_p_value']) and r['mult_p_value'] < 0.05:
            pass
            ws.cell(row=rn,column=ci)

    sr = rn + 2
    ws.cell(row=sr,column=1,value="Summary")
    valid = res_df.dropna(subset=['direct_slope'])
    for i, (l, v) in enumerate([
        ("National multiplier", f"{NAT_MULT:.3f}"),
        ("DC industry share", f"{100/NAT_MULT:.2f}%"),
        ("Median local_share", f"{res_df['local_share'].median():.1%}"),
        ("",""),
        ("Valid regressions", f"{len(valid)} / {len(res_df)}"),
        (f"Median β cleaned direct (jobs/{x_unit})", f"{valid['direct_slope'].median():,.1f}"),
        (f"Median β cleaned×mult (jobs/{x_unit})", f"{valid['mult_slope'].median():,.1f}"),
        ("p<0.05 (cleaned direct)", f"{(valid['direct_p_value']<0.05).sum()} / {len(valid)}"),
        ("p<0.05 (cleaned×mult)", f"{(valid['mult_p_value']<0.05).sum()} / {len(valid)}"),
        ("",""),
        ("Method", "cleaned_direct = NAICS_5182 × local_share"),
        ("", "cleaned_multiplied = cleaned_direct × local_multiplier"),
        ("local_share", "PwC_direct / (PwC_direct + spillover/nat_mult)"),
    ]):
        ws.cell(row=sr+1+i,column=1,value=l)
        ws.cell(row=sr+1+i,column=3,value=v)

    for col, w in {'A':7,'B':22,'C':5,'D':8,'E':8,'F':9,'G':9,
                   'H':13,'I':10,'J':13,'K':13,'L':12,'M':8,'N':9,
                   'O':13,'P':10,'Q':13,'R':13,'S':12,'T':8,'U':9}.items():
        pass

def write_panel_sheet(wb, panel_df):
    ws = wb.create_sheet("Panel Data")
    pass
    ws.freeze_panes = "A2"
    pass

    headers = ['State', 'State Name', 'Year', 'NAICS 5182\n(observed)',
               'DC Count', 'DC Cap\n(GW)', 'Local\nShare', 'Local\nMultiplier',
               'Cleaned\nDirect', 'Cleaned\n×Multiplier', 'Raw\nMultiplied']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        pass
        pass
        pass
    ws.row_dimensions[1].height = 35

    fmts = [None, None, None, '#,##0.0', '#,##0', '0.000', '0.0%', '0.00', '#,##0.0', '#,##0', '#,##0']
    for idx, (_, r) in enumerate(panel_df.sort_values(['state_abbr','year']).iterrows()):
        rn = idx + 2
        vals = [r['state_abbr'], r['state_name'], int(r['year']), r['emp_5182'],
                int(r['dc_count']), r['dc_capacity_GW'], r['local_share'], r['local_multiplier'],
                r['emp_cleaned_direct'], r['emp_cleaned_multiplied'], r['emp_raw_multiplied']]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=rn, column=ci, value=v)
            pass
            pass

    for ci, w in enumerate([7,22,7,12,9,10,8,8,12,12,12], 1):
        pass

# Build workbook
wb = Workbook()
wb.remove(wb.active)

write_raw_sheet(wb, "Raw OLS by Capacity", "4472C4", res_raw_cap,
                "DC Capacity", "GW", "0.000", "4472C4", "2F5496", "D6E4F0", "D6E4F0", "#,##0")
write_panel_sheet(wb, panel)
write_cleaned_sheet(wb, "Cleaned OLS by Capacity", "7030A0", res_cln_cap,
                    "DC Capacity", "GW", "0.000", "7030A0", "5B259F", "#,##0")

# ============================================================
# 6. Panel Fixed Effects: Emp_it = α_i + γ_t + β × X_it + ε_it
# ============================================================
def run_panel_fe(df, x_col, y_col, entity_col='state_abbr', time_col='year'):
    pdf = df[[entity_col, time_col, x_col, y_col]].dropna().copy()
    pdf = pdf.set_index([entity_col, time_col])
    mod = PanelOLS(pdf[y_col], pdf[[x_col]], entity_effects=True, time_effects=True)
    return mod.fit(cov_type='clustered', cluster_entity=True)

def cgm_se(res_fe, x_col):
    """
    Cameron-Gelbach-Miller finite-sample-corrected clustered SE and t(G-1)
    inference -- the convention SI section 6 prints.  linearmodels reports the
    uncorrected sandwich; the two differ by exactly
        sqrt( G/(G-1) * (N-1)/N ).
    Both are written to the output so the conventions cannot silently diverge
    across script boundaries.
    """
    n = int(res_fe.nobs)
    g = int(res_fe.entity_info['total'])
    factor = np.sqrt(g / (g - 1) * (n - 1) / n)
    se = float(res_fe.std_errors[x_col]) * factor
    beta = float(res_fe.params[x_col])
    t = beta / se
    p = 2 * stats.t.sf(abs(t), g - 1)
    return se, t, p, factor, n, g

fe_specs = []
for y_col, y_label in [
    ('emp_5182', 'Raw NAICS 5182'),
    ('emp_cleaned_direct', 'Cleaned Direct'),
    ('emp_raw_multiplied', 'Raw × Multiplier+Spillover'),
    ('emp_state_total', 'State total (local impact + spillover)'),
    ('emp_cleaned_multiplied', 'Cleaned × Local Multiplier'),
]:
    res = run_panel_fe(panel, 'dc_capacity_GW', y_col)
    se_c, t_c, p_c, fac, n_c, g_c = cgm_se(res, 'dc_capacity_GW')
    fe_specs.append({'x_col': 'dc_capacity_GW', 'x_label': 'Capacity (GW)',
                     'y_col': y_col, 'y_label': y_label, 'result': res,
                     'se_cgm': se_c, 't_cgm': t_c, 'p_cgm': p_c,
                     'cgm_factor': fac, 'n_obs': n_c, 'n_units': g_c})

# The same four specifications, but with the UNBALANCED national total in the
# spec-2 spillover term.  Reported as a diagnostic only: it is the size of the
# artefact an unbalanced aggregate would introduce into this table.
_panel_unbal = panel.copy()
_panel_unbal['emp_raw_multiplied'] = _panel_unbal.apply(
    lambda r: (r['emp_5182'] * local_mult_dict[r['state_name']]
               + (direct_by_year_unbalanced[r['year']] - r['emp_5182'])
               * spillover_rate_dict[r['state_name']]), axis=1)
_res_unbal = run_panel_fe(_panel_unbal, 'dc_capacity_GW', 'emp_raw_multiplied')
_se_unbal, _, _, _, _, _ = cgm_se(_res_unbal, 'dc_capacity_GW')
print(f"\n[diagnostic] Raw x mult + spillover, UNBALANCED national total: "
      f"beta = {float(_res_unbal.params['dc_capacity_GW']):,.2f} "
      f"(se_lm {float(_res_unbal.std_errors['dc_capacity_GW']):,.2f}, "
      f"se_cgm {_se_unbal:,.2f})")
print(f"[diagnostic] Raw x mult + spillover, BALANCED national total  : "
      f"beta = {float(fe_specs[2]['result'].params['dc_capacity_GW']):,.2f} "
      f"(se_lm {float(fe_specs[2]['result'].std_errors['dc_capacity_GW']):,.2f}, "
      f"se_cgm {fe_specs[2]['se_cgm']:,.2f})")

def write_panel_fe_sheet(wb):
    ws = wb.create_sheet("Panel FE")
    pass
    ws.freeze_panes = "A3"
    pass

    ws.merge_cells('A1:K1')
    c = ws['A1']
    c.value = "Employment — Panel Fixed Effects: Emp_it = α_i + γ_t + β × X_it + ε_it"
    pass
    pass

    headers = ['Specification (Y)', 'X Variable', 'β', 'SE\n(clustered,\nlinearmodels)',
               'CI95\nLower', 'CI95\nUpper', 'p-value', 'R² within', 'N obs / Units',
               'SE\n(CGM-corrected,\nSI convention)', 'p-value\nt(G−1)']
    pass
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        pass
        pass
        pass
    ws.row_dimensions[2].height = 35

    for idx, sp in enumerate(fe_specs):
        rn = idx + 3
        res_fe = sp['result']
        x = sp['x_col']
        beta = res_fe.params[x]
        se = res_fe.std_errors[x]
        pval = res_fe.pvalues[x]

        vals = [sp['y_label'], sp['x_label'],
                beta, se, beta - 1.96 * se, beta + 1.96 * se,
                pval, res_fe.rsquared_within,
                f"{res_fe.nobs} / {res_fe.entity_info['total']}",
                sp['se_cgm'], sp['p_cgm']]
        fmts = [None, None, '#,##0.0', '#,##0.0', '#,##0.0', '#,##0.0', '0.0000',
                '0.000', None, '#,##0.0', '0.0000']

        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=rn, column=ci, value=v)
            pass
            if fmt:
                pass
        if pval < 0.05:
            for ci in [3, 4, 5, 6, 7]:
                ws.cell(row=rn, column=ci)

    sr = len(fe_specs) + 4
    ws.cell(row=sr, column=1, value="Notes")
    for i, (l, v) in enumerate([
        ("Model", "Emp_it = α_i + γ_t + β × X_it + ε_it"),
        ("Fixed Effects", "Unit (entity) + Year (time)"),
        ("SE", "Clustered by unit; both the linearmodels sandwich and the"),
        ("", "Cameron-Gelbach-Miller finite-sample correction are reported"),
        ("Universe", UNIVERSE_LABEL + f" ({N_ANALYSIS_UNITS} units)"),
        ("Period", "2016–2024 (9 years); Michigan absent 2022–2024"),
        ("Spec-2 base", "national NAICS 5182 total summed over units observed"),
        ("", "in every year of 2016–2024 (balanced), not over all rows"),
        ("Interpretation", "β = avg within-unit employment change per unit X,"),
        ("", "controlling for unit-level and time-level unobservables"),
    ]):
        ws.cell(row=sr + 1 + i, column=1, value=l)
        ws.cell(row=sr + 1 + i, column=3, value=v)

    for ci, w in enumerate([28, 16, 12, 14, 12, 12, 10, 10, 14, 16, 11], 1):
        pass

write_panel_fe_sheet(wb)

wb.save(_DIR / '02_ols_employment_dc_capacity.xlsx')
print("Saved employment/02_ols_employment_dc_capacity.xlsx")

for label, res in [("Raw by Capacity", res_raw_cap),
                   ("Cleaned by Capacity", res_cln_cap)]:
    v = res.dropna(subset=['direct_slope'])
    print(f"\n=== {label} ===")
    print(f"  Valid: {len(v)}/{len(res)}")
    print(f"  Median β direct: {v['direct_slope'].median():,.1f}")
    print(f"  Median β multiplied: {v['mult_slope'].median():,.1f}")
    print(f"  p<0.05: {(v['direct_p_value']<0.05).sum()}/{len(v)} direct, "
          f"{(v['mult_p_value']<0.05).sum()}/{len(v)} mult")
    print(f"  p<0.05 and positive (direct): "
          f"{((v['direct_p_value']<0.05) & (v['direct_slope']>0)).sum()}/{len(v)}")

print("\n=== Panel FE ===")
_fe_rows = []
for sp in fe_specs:
    res_fe = sp['result']
    x = sp['x_col']
    beta = float(res_fe.params[x])
    p = float(res_fe.pvalues[x])
    sig = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else ''
    print(f"  {sp['y_label']:35s} ~ {sp['x_label']:15s}  β={beta:>10,.1f}  "
          f"se_lm={float(res_fe.std_errors[x]):>8,.1f}  se_cgm={sp['se_cgm']:>8,.1f}  "
          f"p={p:.4f}{sig}  p_cgm={sp['p_cgm']:.4f}  N={int(res_fe.nobs)}/"
          f"{int(res_fe.entity_info['total'])}")
    _fe_rows.append({
        'script': '09_ols_employment_dc.py', 'spec': sp['y_label'],
        'beta': beta, 'se_linearmodels': float(res_fe.std_errors[x]),
        'p_linearmodels': p, 'se_cgm': sp['se_cgm'], 'p_cgm_t_Gm1': sp['p_cgm'],
        'r2_within': float(res_fe.rsquared_within),
        'n_obs': int(res_fe.nobs), 'n_units': int(res_fe.entity_info['total']),
    })

_fe_rows.append({
    'script': '09_ols_employment_dc.py',
    'spec': 'Raw × Multiplier+Spillover [DIAGNOSTIC: unbalanced national total]',
    'beta': float(_res_unbal.params['dc_capacity_GW']),
    'se_linearmodels': float(_res_unbal.std_errors['dc_capacity_GW']),
    'p_linearmodels': float(_res_unbal.pvalues['dc_capacity_GW']),
    'se_cgm': _se_unbal, 'p_cgm_t_Gm1': np.nan,
    'r2_within': float(_res_unbal.rsquared_within),
    'n_obs': int(_res_unbal.nobs), 'n_units': int(_res_unbal.entity_info['total']),
})

_out = _DIR / '../results/r6_employment' / 'panel_fe_employment.csv'
pd.DataFrame(_fe_rows).to_csv(_out, index=False)
print(f"\nSaved {_out}")

_ols_out = _DIR / '../results/r6_employment' / 'state_ols_employment.csv'
res_cln_cap.assign(spec='cleaned_direct').to_csv(_ols_out, index=False)
print(f"Saved {_ols_out}")
