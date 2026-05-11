"""
Intra-industry substitution analysis: NAICS 5182 vs 517 vs 51
For states with DC capacity > 1GW (2024) + national total.

Compares employment changes (2016→2024) across:
  - NAICS 5182: Data Processing, Hosting & Related
  - NAICS 51:  Information Sector (parent)
  - NAICS 517: Telecommunications

Output: 03_substitution_analysis.xlsx
  Sheet 1: Substitution Summary (side-by-side changes)
  Sheet 2: Time Series (annual employment 2016-2024)
  Sheet 3: 5182 vs 517 Comparison (net analysis & substitution ratio)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 1. Load data
# ============================================================
_DIR = Path(__file__).resolve().parent / "employment"

qwi = pd.read_csv(_DIR / 'qwi_all_naics_annual.csv', dtype={'naics': str})
dc = pd.read_csv(_DIR / 'dc_facilities_by_state_year.csv')

STATE_ABBR_MAP = {
    'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA',
    'Colorado':'CO','Connecticut':'CT','Delaware':'DE','District of Columbia':'DC',
    'Florida':'FL','Georgia':'GA','Hawaii':'HI','Idaho':'ID','Illinois':'IL',
    'Indiana':'IN','Iowa':'IA','Kansas':'KS','Kentucky':'KY','Louisiana':'LA',
    'Maine':'ME','Maryland':'MD','Massachusetts':'MA','Michigan':'MI','Minnesota':'MN',
    'Mississippi':'MS','Missouri':'MO','Montana':'MT','Nebraska':'NE','Nevada':'NV',
    'New Hampshire':'NH','New Jersey':'NJ','New Mexico':'NM','New York':'NY',
    'North Carolina':'NC','North Dakota':'ND','Ohio':'OH','Oklahoma':'OK','Oregon':'OR',
    'Pennsylvania':'PA','Rhode Island':'RI','South Carolina':'SC','South Dakota':'SD',
    'Tennessee':'TN','Texas':'TX','Utah':'UT','Vermont':'VT','Virginia':'VA',
    'Washington':'WA','West Virginia':'WV','Wisconsin':'WI','Wyoming':'WY'
}
ABBR_TO_NAME = {v: k for k, v in STATE_ABBR_MAP.items()}
YEARS = list(range(2016, 2025))

# ============================================================
# 2. Get 2024 DC capacity and filter states > 1GW
# ============================================================
dc_2024 = dc[dc['state_abbr'] != 'US'][['state_abbr', 'MW_2024']].copy()
dc_2024['cap_GW'] = dc_2024['MW_2024'] / 1000
dc_2024 = dc_2024[dc_2024['cap_GW'] >= 1.0].sort_values('cap_GW', ascending=False)
big_states = list(dc_2024['state_abbr'])

# National total capacity
nat_cap = dc[dc['state_abbr'] == 'US']['MW_2024'].values[0] / 1000

# ============================================================
# 3. Build employment pivots for 3 NAICS codes
# ============================================================
qwi['state_abbr'] = qwi['state_name'].map(STATE_ABBR_MAP)

pivots = {}
for naics in ['5182', '51', '517']:
    q = qwi[qwi['naics'] == naics].copy()
    piv = q.pivot_table(index='state_abbr', columns='year', values='Emp_annual_avg')
    pivots[naics] = piv

# National totals
nat_totals = {}
for naics in ['5182', '51', '517']:
    q = qwi[qwi['naics'] == naics]
    nat_totals[naics] = q.groupby('year')['Emp_annual_avg'].sum()

# ============================================================
# 4. Build summary table
# ============================================================
rows = []
for abbr in big_states:
    st_name = ABBR_TO_NAME.get(abbr, abbr)
    cap = dc_2024[dc_2024['state_abbr'] == abbr]['cap_GW'].values[0]
    row = {'state_abbr': abbr, 'state_name': st_name, 'cap_GW': cap}

    for naics, prefix in [('5182', 'dc'), ('51', 'info'), ('517', 'tel')]:
        if abbr in pivots[naics].index:
            e16 = pivots[naics].loc[abbr, 2016] if 2016 in pivots[naics].columns else np.nan
            e24 = pivots[naics].loc[abbr, 2024] if 2024 in pivots[naics].columns else np.nan
        else:
            e16 = e24 = np.nan
        delta = e24 - e16 if not (np.isnan(e16) or np.isnan(e24)) else np.nan
        pct = delta / e16 if (e16 and not np.isnan(delta)) else np.nan
        row[f'{prefix}_2016'] = e16
        row[f'{prefix}_2024'] = e24
        row[f'{prefix}_delta'] = delta
        row[f'{prefix}_pct'] = pct
    rows.append(row)

# National row
nat_row = {'state_abbr': 'US', 'state_name': 'National', 'cap_GW': nat_cap}
for naics, prefix in [('5182', 'dc'), ('51', 'info'), ('517', 'tel')]:
    e16 = nat_totals[naics].get(2016, np.nan)
    e24 = nat_totals[naics].get(2024, np.nan)
    delta = e24 - e16
    nat_row[f'{prefix}_2016'] = e16
    nat_row[f'{prefix}_2024'] = e24
    nat_row[f'{prefix}_delta'] = delta
    nat_row[f'{prefix}_pct'] = delta / e16 if e16 else np.nan
rows.append(nat_row)

summary = pd.DataFrame(rows)

# Interpretation columns
def interpret_pattern(r):
    dc_d = r['dc_delta']
    info_d = r['info_delta']
    if pd.isna(dc_d) or pd.isna(info_d):
        return ''
    if dc_d > 0 and info_d < 0:
        return 'Substitution'
    elif dc_d > 0 and info_d > 0:
        if dc_d > info_d:
            return 'Substitution'
        else:
            return 'Net new jobs'
    elif dc_d < 0:
        return 'DC declining'
    return 'Net new jobs'

def interpret_tel(r):
    tel_pct = r['tel_pct']
    if pd.isna(tel_pct):
        return ''
    if tel_pct < -0.1:
        return 'Large decline'
    elif tel_pct < 0:
        return 'Decline'
    return 'Stable/Growth'

summary['pattern'] = summary.apply(interpret_pattern, axis=1)
summary['tel_interp'] = summary.apply(interpret_tel, axis=1)

# ============================================================
# 5. Write Excel
# ============================================================
thin = Side(style='thin', color='000000')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)

# --- Sheet 1: Substitution Summary ---
ws = wb.create_sheet("Substitution Summary")
ws.sheet_properties.tabColor = "4472C4"
ws.freeze_panes = "A3"
hf = Font(bold=True, color="FFFFFF", size=11)

groups = [
    ('A1:C1', 'State', '4472C4'),
    ('D1:G1', 'NAICS 5182 — Data Processing & Hosting', '2F5496'),
    ('H1:K1', 'NAICS 51 — Information Sector', '548235'),
    ('L1:O1', 'NAICS 517 — Telecommunications', 'BF8F00'),
    ('P1:Q1', 'Interpretation', '7030A0'),
]
for rng, title, color in groups:
    ws.merge_cells(rng)
    c = ws[rng.split(':')[0]]
    c.value = title; c.font = hf
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal='center')

headers = ['State', 'State Name', 'DC Cap\n2024 (GW)',
           'Emp\n2016', 'Emp\n2024', 'Δ Change', '% Change',
           'Emp\n2016', 'Emp\n2024', 'Δ Change', '% Change',
           'Emp\n2016', 'Emp\n2024', 'Δ Change', '% Change',
           '5182 vs 51\nPattern', '517\nDecline']
fills = ['D6E4F0'] * 3 + ['D6E4F0'] * 4 + ['E2EFDA'] * 4 + ['FFF2CC'] * 4 + ['E8D5F5'] * 2
for ci, (h, fc) in enumerate(zip(headers, fills), 1):
    cell = ws.cell(row=2, column=ci, value=h)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
    cell.fill = PatternFill("solid", fgColor=fc)
ws.row_dimensions[2].height = 35

fmts = [None, None, '0.000',
        '#,##0', '#,##0', '+#,##0;-#,##0', '+0.0%;-0.0%',
        '#,##0', '#,##0', '+#,##0;-#,##0', '+0.0%;-0.0%',
        '#,##0', '#,##0', '+#,##0;-#,##0', '+0.0%;-0.0%',
        None, None]
keys = ['state_abbr', 'state_name', 'cap_GW',
        'dc_2016', 'dc_2024', 'dc_delta', 'dc_pct',
        'info_2016', 'info_2024', 'info_delta', 'info_pct',
        'tel_2016', 'tel_2024', 'tel_delta', 'tel_pct',
        'pattern', 'tel_interp']

for idx, (_, r) in enumerate(summary.iterrows()):
    rn = idx + 3
    for ci, (k, fmt) in enumerate(zip(keys, fmts), 1):
        v = r[k]
        if pd.isna(v):
            v = None
        cell = ws.cell(row=rn, column=ci, value=v)
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt

for ci, w in enumerate([5, 16, 10, 10, 10, 10, 9, 10, 10, 10, 9, 10, 10, 10, 9, 14, 12], 1):
    ws.column_dimensions[chr(64 + ci) if ci <= 26 else chr(64 + ci - 26)].width = w

# --- Sheet 2: Time Series ---
ws2 = wb.create_sheet("Time Series")
ws2.sheet_properties.tabColor = "548235"
ws2.freeze_panes = "A3"

col_offset = 0
for naics, title, color in [('5182', 'NAICS 5182 — Data Processing & Hosting', '2F5496'),
                              ('51', 'NAICS 51 — Information Sector', '548235'),
                              ('517', 'NAICS 517 — Telecommunications', 'BF8F00')]:
    start_col = col_offset + 1
    end_col = col_offset + 2 + len(YEARS)

    ws2.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    c = ws2.cell(row=1, column=start_col, value=title)
    c.font = hf; c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal='center')

    for ci, h in enumerate(['State', 'State Name'] + [str(y) for y in YEARS], start_col):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True, size=10)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    for idx, abbr in enumerate(big_states + ['US']):
        rn = idx + 3
        ws2.cell(row=rn, column=start_col, value=abbr).border = BORDER
        ws2.cell(row=rn, column=start_col + 1, value=ABBR_TO_NAME.get(abbr, 'National')).border = BORDER

        if abbr == 'US':
            for yi, yr in enumerate(YEARS):
                v = nat_totals[naics].get(yr, None)
                cell = ws2.cell(row=rn, column=start_col + 2 + yi, value=v)
                cell.border = BORDER; cell.number_format = '#,##0'
        elif abbr in pivots[naics].index:
            for yi, yr in enumerate(YEARS):
                v = pivots[naics].loc[abbr, yr] if yr in pivots[naics].columns else None
                cell = ws2.cell(row=rn, column=start_col + 2 + yi, value=v)
                cell.border = BORDER; cell.number_format = '#,##0'

    col_offset = end_col + 1

# --- Sheet 3: 5182 vs 517 Comparison ---
ws3 = wb.create_sheet("5182 vs 517 Comparison")
ws3.sheet_properties.tabColor = "7030A0"
ws3.freeze_panes = "A3"

groups3 = [
    ('A1:C1', 'State', '4472C4'),
    ('D1:E1', 'NAICS 5182 (DC)', '2F5496'),
    ('F1:G1', 'NAICS 517 (Telecom)', 'BF8F00'),
    ('H1:I1', 'NAICS 51 (Info)', '548235'),
    ('J1:K1', 'Net Analysis', 'C00000'),
]
for rng, title, color in groups3:
    ws3.merge_cells(rng)
    c = ws3[rng.split(':')[0]]
    c.value = title; c.font = hf
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal='center')

headers3 = ['State', 'State Name', 'DC Cap\n(GW)',
            'Δ 5182', '% Chg', 'Δ 517', '% Chg', 'Δ 51', '% Chg',
            '5182+517\nNet', 'Substitution\nRatio']
for ci, h in enumerate(headers3, 1):
    cell = ws3.cell(row=2, column=ci, value=h)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
ws3.row_dimensions[2].height = 35

fmts3 = [None, None, '0.000',
         '+#,##0;-#,##0', '+0.0%;-0.0%',
         '+#,##0;-#,##0', '+0.0%;-0.0%',
         '+#,##0;-#,##0', '+0.0%;-0.0%',
         '+#,##0;-#,##0', '0.00']

for idx, (_, r) in enumerate(summary.iterrows()):
    rn = idx + 3
    net = (r['dc_delta'] or 0) + (r['tel_delta'] or 0)
    sub_ratio = abs(r['dc_delta'] / r['tel_delta']) if (r['tel_delta'] and r['tel_delta'] != 0) else np.nan

    vals = [r['state_abbr'], r['state_name'], r['cap_GW'],
            r['dc_delta'], r['dc_pct'], r['tel_delta'], r['tel_pct'],
            r['info_delta'], r['info_pct'], net, sub_ratio]
    for ci, (v, fmt) in enumerate(zip(vals, fmts3), 1):
        if pd.isna(v):
            v = None
        cell = ws3.cell(row=rn, column=ci, value=v)
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt

for ci, w in enumerate([5, 16, 8, 10, 8, 10, 8, 10, 8, 10, 10], 1):
    ws3.column_dimensions[chr(64 + ci)].width = w

wb.save(_DIR / '03_substitution_analysis.xlsx')
print(f"Saved 03_substitution_analysis.xlsx")
print(f"States > 1GW: {len(big_states)} + national = {len(big_states)+1} rows")
