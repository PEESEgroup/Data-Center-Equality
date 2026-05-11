"""
Incremental labor income (2020→2024) vs cumulative state fiscal incentives (2020-2024).

Method:
  Δ emp = NAICS 5182 (2024) − NAICS 5182 (2020)
  Δ cleaned = Δ raw × local_share
  Δ direct LI = Δ cleaned × annual earnings (QWI EarnBeg 2024 × 12)
  Δ total LI = Δ direct LI × li_multiplier_local (PwC, no spillover)
  Subsidy = cumulative state incentives 2020–2024

Output: 05_labor_income_vs_subsidy.xlsx
  Sheet 1: Incremental LI vs Subsidy (states with subsidy > 0)
  Sheet 2: All States (sorted by Δ total LI)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 1. Load data
# ============================================================
_DIR = Path(__file__).resolve().parent / "employment"

qwi = pd.read_csv(_DIR / 'qwi_all_naics_annual.csv', dtype={'naics': str})
mult = pd.read_csv(_DIR / 'pwc_multipliers.csv')
sub = pd.read_excel(_DIR / 'tax' / 'state_subsidy_by_year_million_wide.xlsx')

STATE_ABBR_MAP = {
    'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA',
    'Colorado':'CO','Connecticut':'CT','Delaware':'DE','District of Columbia':'DC',
    'Florida':'FL','Georgia':'GA','Hawaii':'HI','Idaho':'ID','Illinois':'IL',
    'Indiana':'IN','Iowa':'IA','Kansas':'KS','Kentucky':'KY','Louisiana':'LA',
    'Maine':'ME','Maryland':'MD','Massachusetts':'MA','Michigan':'MI','Minnesota':'MN',
    'Mississippi':'MS','Missouri':'MO','Montana':'MT','Nebraska':'NE','Nevada':'NV',
    'New Hampshire':'NH','New Jersey':'NJ','New Mexico':'NM','New York':'NY',
    'North Carolina':'NC','North Dakota':'ND','Ohio':'OH','Oklahoma':'OK','Oregon':'OR',
    'Pennsylvania':'PA','Rhode Island':'RI','South Carolina':'SC','South Dakota':'SD',
    'Tennessee':'TN','Texas':'TX','Utah':'UT','Vermont':'VT','Virginia':'VA',
    'Washington':'WA','West Virginia':'WV','Wisconsin':'WI','Wyoming':'WY'
}

# ============================================================
# 2. Spillover decomposition → local_share
# ============================================================
NAT_MULT = mult['emp_total_no_spillover_avg'].sum() / mult['emp_direct_avg'].sum()
mult['spillover_dc'] = mult['emp_spillover_avg'] / NAT_MULT
mult['local_share'] = mult['emp_direct_avg'] / (mult['emp_direct_avg'] + mult['spillover_dc'])

local_share_dict = dict(zip(mult['state'], mult['local_share']))
li_mult_dict = dict(zip(mult['state'], mult['li_multiplier_local']))

# ============================================================
# 3. Subsidy: cumulative 2020-2024
# ============================================================
sub['subsidy_total_M'] = sub[[2020, 2021, 2022, 2023, 2024]].sum(axis=1)
sub_dict = dict(zip(sub['State'], sub['subsidy_total_M']))

# ============================================================
# 4. Build incremental table: Δ employment 2020→2024
# ============================================================
q5182 = qwi[qwi['naics'] == '5182'].copy()

rows = []
for st_name, abbr in STATE_ABBR_MAP.items():
    lookup = st_name if st_name != 'District of Columbia' else 'DC'
    e20 = q5182[(q5182['state_name'] == lookup) & (q5182['year'] == 2020)]
    e24 = q5182[(q5182['state_name'] == lookup) & (q5182['year'] == 2024)]
    if len(e20) == 0 or len(e24) == 0:
        continue

    emp_2020 = e20['Emp_annual_avg'].values[0]
    emp_2024 = e24['Emp_annual_avg'].values[0]
    earn_2024 = e24['EarnBeg_annual_avg'].values[0]

    delta_raw = emp_2024 - emp_2020
    ls = local_share_dict.get(st_name, np.nan)
    lm = li_mult_dict.get(st_name, np.nan)

    delta_cleaned = delta_raw * ls
    annual_earn = earn_2024 * 12
    delta_direct_li = delta_cleaned * annual_earn / 1e6
    delta_total_li = delta_direct_li * lm

    sub_total = sub_dict.get(st_name, 0)
    has_sub = sub_total > 0
    ratio = delta_total_li / sub_total if sub_total > 0 else np.nan
    net = delta_total_li - sub_total if sub_total > 0 else np.nan

    rows.append({
        'state_abbr': abbr, 'state_name': st_name,
        'emp_2020': emp_2020, 'emp_2024': emp_2024,
        'delta_raw': delta_raw, 'local_share': ls,
        'delta_cleaned': delta_cleaned,
        'annual_earnings': annual_earn, 'li_multiplier': lm,
        'delta_direct_li_M': delta_direct_li, 'delta_total_li_M': delta_total_li,
        'subsidy_total_M': sub_total if has_sub else np.nan,
        'has_subsidy': has_sub,
        'ratio': ratio, 'net_M': net,
    })

df = pd.DataFrame(rows)

# ============================================================
# 5. Write Excel
# ============================================================
thin = Side(style='thin', color='000000')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
POS_FILL = PatternFill("solid", fgColor="C6EFCE")
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")

wb = Workbook()
wb.remove(wb.active)

# --- Sheet 1: States with subsidy ---
has_sub_df = df[df['has_subsidy']].sort_values('subsidy_total_M', ascending=False)

ws = wb.create_sheet("Incremental LI vs Subsidy")
ws.sheet_properties.tabColor = "4472C4"
ws.freeze_panes = "A3"
hf = Font(bold=True, color="FFFFFF", size=11)

groups = [
    ('A1:B1', 'State Info', '4472C4'),
    ('C1:G1', 'Employment Change (2020→2024)', '2F5496'),
    ('H1:I1', 'Earnings', '548235'),
    ('J1:K1', 'Δ Labor Income ($M)', '7030A0'),
    ('L1:L1', 'Subsidy ($M)', 'BF8F00'),
    ('M1:N1', 'LI vs Subsidy', 'C00000'),
]
for rng, title, color in groups:
    ws.merge_cells(rng)
    c = ws[rng.split(':')[0]]
    c.value = title; c.font = hf
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal='center')

cols = [
    ('A', 'State'), ('B', 'State Name'),
    ('C', 'NAICS 5182\n2020'), ('D', 'NAICS 5182\n2024'), ('E', 'Δ Raw'),
    ('F', 'Local\nShare'), ('G', 'Δ Cleaned'),
    ('H', 'Annual\nEarnings'), ('I', 'LI\nMultiplier'),
    ('J', 'Δ Direct\nLI ($M)'), ('K', 'Δ Total\nLI ($M)'),
    ('L', 'Subsidy\n20-24 ($M)'),
    ('M', 'ΔLI/Sub\nRatio'), ('N', 'Net ΔLI\n($M)'),
]
sub_fills = {
    'AB': 'D6E4F0', 'CDEFG': 'D6E4F0', 'HI': 'E2EFDA',
    'JK': 'E8D5F5', 'L': 'FFF2CC', 'MN': 'FCE4D6'
}
for cl, title in cols:
    cell = ws[f'{cl}2']
    cell.value = title; cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
    for chars, color in sub_fills.items():
        if cl in chars:
            cell.fill = PatternFill("solid", fgColor=color)
ws.row_dimensions[2].height = 38

fmts = [None, None, '#,##0', '#,##0', '+#,##0;-#,##0', '0.0%', '+#,##0.0;-#,##0.0',
        '$#,##0', '0.00', '+#,##0.0;-#,##0.0', '+#,##0.0;-#,##0.0', '#,##0.0',
        '+#,##0.0x;-#,##0.0x', '+#,##0.0;-#,##0.0']
keys = ['state_abbr', 'state_name', 'emp_2020', 'emp_2024', 'delta_raw',
        'local_share', 'delta_cleaned', 'annual_earnings', 'li_multiplier',
        'delta_direct_li_M', 'delta_total_li_M', 'subsidy_total_M', 'ratio', 'net_M']

for idx, (_, r) in enumerate(has_sub_df.iterrows()):
    rn = idx + 3
    for ci, (key, fmt) in enumerate(zip(keys, fmts), 1):
        v = r[key]
        if pd.isna(v):
            v = None
        cell = ws.cell(row=rn, column=ci, value=v)
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
    # Color net & ratio
    if r['net_M'] is not None and not np.isnan(r['net_M']):
        fill = POS_FILL if r['net_M'] > 0 else NEG_FILL
        ws.cell(row=rn, column=14).fill = fill
        ws.cell(row=rn, column=13).fill = fill
    # Color delta
    fill_d = POS_FILL if r['delta_cleaned'] > 0 else NEG_FILL
    for ci in [5, 7]:
        ws.cell(row=rn, column=ci).fill = fill_d

# Summary
sr = len(has_sub_df) + 4
ws.cell(row=sr, column=1, value="Summary").font = Font(bold=True, size=11)
pos_net = has_sub_df[has_sub_df['net_M'] > 0]
neg_net = has_sub_df[has_sub_df['net_M'] <= 0]

summaries = [
    ("Period", "2020 → 2024"),
    ("States with subsidy", f"{len(has_sub_df)}"),
    ("", ""),
    ("Δ Employment positive", f"{(has_sub_df['delta_raw'] > 0).sum()} / {len(has_sub_df)}"),
    ("Δ Employment negative", f"{(has_sub_df['delta_raw'] <= 0).sum()} / {len(has_sub_df)}"),
    ("", ""),
    ("Net ΔLI > Subsidy", f"{len(pos_net)} states"),
    ("Net ΔLI < Subsidy", f"{len(neg_net)} states"),
    ("", ""),
    ("Total Δ LI (subsidy states)", f"${has_sub_df['delta_total_li_M'].sum():+,.1f}M"),
    ("Total Subsidy 2020-2024", f"${has_sub_df['subsidy_total_M'].sum():,.1f}M"),
    ("Aggregate ratio", f"{has_sub_df['delta_total_li_M'].sum() / has_sub_df['subsidy_total_M'].sum():.1f}x"),
    ("", ""),
    ("Method", "Δ emp = NAICS 5182 (2024) − NAICS 5182 (2020)"),
    ("", "Δ cleaned = Δ raw × local_share"),
    ("", "Δ direct LI = Δ cleaned × annual earnings (QWI 2024)"),
    ("", "Δ total LI = Δ direct LI × li_multiplier_local"),
    ("", "Subsidy = cumulative 2020–2024"),
]
for i, (l, v) in enumerate(summaries):
    ws.cell(row=sr + 1 + i, column=1, value=l).font = Font(bold=True if l else False, size=10)
    ws.cell(row=sr + 1 + i, column=3, value=v)

widths = {'A': 7, 'B': 20, 'C': 11, 'D': 11, 'E': 10, 'F': 8, 'G': 11,
          'H': 12, 'I': 7, 'J': 11, 'K': 11, 'L': 12, 'M': 10, 'N': 10}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# --- Sheet 2: All States ---
ws2 = wb.create_sheet("All States")
ws2.sheet_properties.tabColor = "70AD47"
ws2.freeze_panes = "A2"

headers2 = ['State', 'State Name', '2020 Emp', '2024 Emp', 'Δ Raw', 'Local Share',
            'Δ Cleaned', 'Annual Earn', 'LI Mult', 'Δ Direct LI ($M)', 'Δ Total LI ($M)',
            'Has Subsidy', 'Subsidy 20-24 ($M)', 'ΔLI/Subsidy']
for ci, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor="70AD47")
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
ws2.row_dimensions[1].height = 35

fmts2 = [None, None, '#,##0', '#,##0', '+#,##0;-#,##0', '0.0%', '+#,##0.0;-#,##0.0',
         '$#,##0', '0.00', '+#,##0.0;-#,##0.0', '+#,##0.0;-#,##0.0', None, '#,##0.0', '+#,##0.0;-#,##0.0']
keys2 = ['state_abbr', 'state_name', 'emp_2020', 'emp_2024', 'delta_raw',
         'local_share', 'delta_cleaned', 'annual_earnings', 'li_multiplier',
         'delta_direct_li_M', 'delta_total_li_M', 'has_subsidy', 'subsidy_total_M', 'ratio']

all_sorted = df.sort_values('delta_total_li_M', ascending=False)
for idx, (_, r) in enumerate(all_sorted.iterrows()):
    rn = idx + 2
    for ci, (key, fmt) in enumerate(zip(keys2, fmts2), 1):
        v = r[key]
        if pd.isna(v):
            v = None
        if key == 'has_subsidy':
            v = 'Yes' if v else 'No'
        cell = ws2.cell(row=rn, column=ci, value=v)
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
    fill_d = POS_FILL if r['delta_raw'] > 0 else NEG_FILL
    ws2.cell(row=rn, column=5).fill = fill_d

for ci, w in enumerate([7, 20, 10, 10, 9, 8, 10, 11, 7, 11, 11, 8, 11, 10], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

wb.save(_DIR / '05_labor_income_vs_subsidy.xlsx')
print("Saved 05_labor_income_vs_subsidy.xlsx")

# Print summary
print(f"\nNational Δ raw: {df['delta_raw'].sum():+,.0f}")
print(f"National Δ total LI: ${df['delta_total_li_M'].sum():+,.0f}M")
print(f"Subsidy states: {len(has_sub_df)}")
print(f"  Total Δ LI: ${has_sub_df['delta_total_li_M'].sum():+,.0f}M")
print(f"  Total subsidy: ${has_sub_df['subsidy_total_M'].sum():,.0f}M")
print(f"  Ratio: {has_sub_df['delta_total_li_M'].sum() / has_sub_df['subsidy_total_M'].sum():.1f}x")
